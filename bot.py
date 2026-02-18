import logging
import sqlite3
from datetime import datetime, date, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ConversationHandler, MessageHandler, filters, ContextTypes
import os
import csv
import io
import openpyxl
import tempfile
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv

# Загрузка токена
load_dotenv()
BOT_TOKEN = os.getenv('BOT_TOKEN')

# Настройка логирования
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

# Состояния для регистрации
REGISTER_NAME, REGISTER_POSITION, REGISTER_STORE = range(3)

# Состояния для добавления администратора
ADD_ADMIN_ID, ADD_ADMIN_CONFIRM = range(3, 5)

# Состояния для создания должности
CREATE_POSITION_NAME = 5

# Состояния для создания магазина
CREATE_STORE_NAME, CREATE_STORE_ADDRESS = range(6, 8)

# Состояния для удаления
DELETE_SELECT, DELETE_CONFIRM, DELETE_SUPER_CONFIRM = range(8, 11)

# Состояния для заявки на администратора
BECOME_ADMIN_REQUEST, BECOME_ADMIN_CONFIRM = range(11, 13)

# Состояния для назначения супер-администратора
ASSIGN_SUPER_ADMIN_SELECT, ASSIGN_SUPER_ADMIN_CONFIRM = range(13, 15)

# НОВЫЕ СОСТОЯНИЯ ДЛЯ ВЫБОРА ПЕРИОДА
SELECT_PERIOD_START, SELECT_PERIOD_END, SELECT_PERIOD_TYPE = range(15, 18)

# Инициализация базы данных
def init_database():
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    
    # Таблица сотрудников
    cursor.execute('''CREATE TABLE IF NOT EXISTS employees 
                      (user_id INTEGER PRIMARY KEY, 
                       full_name TEXT, 
                       position TEXT, 
                       store TEXT,
                       reg_date TEXT, 
                       is_admin INTEGER DEFAULT 0,
                       is_super_admin INTEGER DEFAULT 0)''')
    
    # Таблица записей табеля
    cursor.execute('''CREATE TABLE IF NOT EXISTS timesheet 
                      (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                       user_id INTEGER, 
                       date TEXT, 
                       status TEXT, 
                       check_in TEXT, 
                       check_out TEXT, 
                       hours REAL, 
                       notes TEXT,
                       confirmed INTEGER DEFAULT 0)''')
    
    # Таблица для запросов на удаление
    cursor.execute('''CREATE TABLE IF NOT EXISTS delete_requests 
                      (id INTEGER PRIMARY KEY AUTOINCREMENT,
                       request_date TEXT,
                       requester_id INTEGER,
                       requester_name TEXT,
                       target_type TEXT,
                       target_id TEXT,
                       target_name TEXT,
                       status TEXT DEFAULT 'pending')''')
    
    # Таблица для заявок на администратора
    cursor.execute('''CREATE TABLE IF NOT EXISTS admin_requests 
                      (id INTEGER PRIMARY KEY AUTOINCREMENT,
                       request_date TEXT,
                       user_id INTEGER,
                       user_name TEXT,
                       user_position TEXT,
                       user_store TEXT,
                       status TEXT DEFAULT 'pending')''')
    
    # Таблица должностей
    cursor.execute('''CREATE TABLE IF NOT EXISTS positions 
                      (id INTEGER PRIMARY KEY AUTOINCREMENT,
                       name TEXT UNIQUE,
                       created_by INTEGER,
                       created_date TEXT)''')
    
    # Таблица магазинов
    cursor.execute('''CREATE TABLE IF NOT EXISTS stores 
                      (id INTEGER PRIMARY KEY AUTOINCREMENT,
                       name TEXT UNIQUE,
                       address TEXT,
                       created_by INTEGER,
                       created_date TEXT)''')
    
    conn.commit()
    conn.close()

# Функции БД для сотрудников
def add_employee(user_id, name, position, store):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM employees')
    count = cursor.fetchone()[0]
    is_admin = 1 if count == 0 else 0
    is_super_admin = 1 if count == 0 else 0
    cursor.execute('INSERT OR REPLACE INTO employees VALUES (?, ?, ?, ?, ?, ?, ?)',
                  (user_id, name, position, store, datetime.now().isoformat(), is_admin, is_super_admin))
    conn.commit()
    conn.close()
    return is_admin, is_super_admin

def get_employee(user_id):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM employees WHERE user_id = ?', (user_id,))
    result = cursor.fetchone()
    conn.close()
    return result

def get_all_employees():
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM employees ORDER BY store, full_name')
    result = cursor.fetchall()
    conn.close()
    return result

def get_employees_by_store(store):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM employees WHERE store = ? ORDER BY full_name', (store,))
    result = cursor.fetchall()
    conn.close()
    return result

def get_all_admins():
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM employees WHERE is_admin = 1 AND is_super_admin = 0 ORDER BY store, full_name')
    result = cursor.fetchall()
    conn.close()
    return result

def get_super_admin():
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT user_id FROM employees WHERE is_super_admin = 1')
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else None

def get_all_super_admins():
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM employees WHERE is_super_admin = 1 ORDER BY full_name')
    result = cursor.fetchall()
    conn.close()
    return result

def is_super_admin(user_id):
    emp = get_employee(user_id)
    return emp and len(emp) > 6 and emp[6] == 1

def add_admin(user_id):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('UPDATE employees SET is_admin = 1 WHERE user_id = ?', (user_id,))
    conn.commit()
    conn.close()

def is_admin(user_id):
    emp = get_employee(user_id)
    return emp and emp[5] == 1

def assign_super_admin(user_id):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('UPDATE employees SET is_super_admin = 1, is_admin = 1 WHERE user_id = ?', (user_id,))
    conn.commit()
    conn.close()

def remove_super_admin(user_id):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('UPDATE employees SET is_super_admin = 0 WHERE user_id = ?', (user_id,))
    conn.commit()
    conn.close()

# Функции для заявок на администратора
def create_admin_request(user_id, user_name, user_position, user_store):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('''INSERT INTO admin_requests 
                      (request_date, user_id, user_name, user_position, user_store, status)
                      VALUES (?, ?, ?, ?, ?, ?)''',
                  (datetime.now().isoformat(), user_id, user_name, user_position, user_store, 'pending'))
    request_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return request_id

def get_pending_admin_requests():
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('''SELECT * FROM admin_requests WHERE status = 'pending' ORDER BY request_date''')
    result = cursor.fetchall()
    conn.close()
    return result

def get_admin_request(request_id):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM admin_requests WHERE id = ?', (request_id,))
    result = cursor.fetchone()
    conn.close()
    return result

def update_admin_request_status(request_id, status):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('UPDATE admin_requests SET status = ? WHERE id = ?', (status, request_id))
    conn.commit()
    conn.close()

def has_pending_admin_request(user_id):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM admin_requests WHERE user_id = ? AND status = "pending"', (user_id,))
    count = cursor.fetchone()[0]
    conn.close()
    return count > 0

# Функции для запросов на удаление
def create_delete_request(requester_id, requester_name, target_type, target_id, target_name):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('''INSERT INTO delete_requests 
                      (request_date, requester_id, requester_name, target_type, target_id, target_name, status)
                      VALUES (?, ?, ?, ?, ?, ?, ?)''',
                  (datetime.now().isoformat(), requester_id, requester_name, target_type, target_id, target_name, 'pending'))
    request_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return request_id

def get_pending_requests():
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('''SELECT * FROM delete_requests WHERE status = 'pending' ORDER BY request_date''')
    result = cursor.fetchall()
    conn.close()
    return result

def get_request(request_id):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM delete_requests WHERE id = ?', (request_id,))
    result = cursor.fetchone()
    conn.close()
    return result

def update_request_status(request_id, status):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('UPDATE delete_requests SET status = ? WHERE id = ?', (status, request_id))
    conn.commit()
    conn.close()

# Функции для удаления сотрудников и магазинов
def delete_employee(user_id):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM timesheet WHERE user_id = ?', (user_id,))
    cursor.execute('DELETE FROM employees WHERE user_id = ?', (user_id,))
    conn.commit()
    conn.close()

def delete_store_with_employees(store_name):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT user_id FROM employees WHERE store = ?', (store_name,))
    employees = cursor.fetchall()
    
    for emp in employees:
        cursor.execute('DELETE FROM timesheet WHERE user_id = ?', (emp[0],))
    
    cursor.execute('DELETE FROM employees WHERE store = ?', (store_name,))
    conn.commit()
    conn.close()

def get_employee_stats(user_id):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM timesheet WHERE user_id = ?', (user_id,))
    entries_count = cursor.fetchone()[0]
    conn.close()
    return entries_count

def get_store_stats(store_name):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM employees WHERE store = ?', (store_name,))
    employees_count = cursor.fetchone()[0]
    
    cursor.execute('''SELECT COUNT(*) FROM timesheet 
                      WHERE user_id IN (SELECT user_id FROM employees WHERE store = ?)''', (store_name,))
    entries_count = cursor.fetchone()[0]
    conn.close()
    return employees_count, entries_count

# Функции для должностей и магазинов
def add_position(name, created_by):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO positions (name, created_by, created_date) VALUES (?, ?, ?)',
                      (name, created_by, datetime.now().isoformat()))
        conn.commit()
        success = True
    except sqlite3.IntegrityError:
        success = False
    conn.close()
    return success

def get_all_positions():
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT name FROM positions ORDER BY name')
    result = cursor.fetchall()
    conn.close()
    return [r[0] for r in result]

def delete_position(name):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM positions WHERE name = ?', (name,))
    conn.commit()
    conn.close()

def add_store(name, address, created_by):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO stores (name, address, created_by, created_date) VALUES (?, ?, ?, ?)',
                      (name, address, created_by, datetime.now().isoformat()))
        conn.commit()
        success = True
    except sqlite3.IntegrityError:
        success = False
    conn.close()
    return success

def get_all_stores():
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT name, address FROM stores ORDER BY name')
    result = cursor.fetchall()
    conn.close()
    return result

def get_store_address(store_name):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('SELECT address FROM stores WHERE name = ?', (store_name,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else ""

def delete_store_from_list(store_name):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM stores WHERE name = ?', (store_name,))
    conn.commit()
    conn.close()

# Функции для табеля
def add_checkin(user_id):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    today = date.today().isoformat()
    now = datetime.now().strftime('%H:%M')
    cursor.execute('INSERT OR REPLACE INTO timesheet (user_id, date, status, check_in) VALUES (?, ?, ?, ?)',
                  (user_id, today, 'working', now))
    conn.commit()
    conn.close()
    return now

def add_checkout(user_id):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    today = date.today().isoformat()
    now = datetime.now().strftime('%H:%M')
    
    cursor.execute('SELECT * FROM timesheet WHERE user_id = ? AND date = ?', (user_id, today))
    entry = cursor.fetchone()
    
    if entry:
        check_in = entry[4]
        check_in_time = datetime.strptime(check_in, '%H:%M')
        check_out_time = datetime.strptime(now, '%H:%M')
        hours = (check_out_time - check_in_time).seconds / 3600
        
        cursor.execute('''UPDATE timesheet SET status = ?, check_out = ?, hours = ? 
                          WHERE user_id = ? AND date = ?''',
                      ('completed', now, hours, user_id, today))
        conn.commit()
        conn.close()
        return now, hours
    conn.close()
    return None, None

def get_timesheet(user_id, days=7):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    start_date = (date.today() - timedelta(days=days)).isoformat()
    cursor.execute('''SELECT * FROM timesheet WHERE user_id = ? AND date >= ? 
                      ORDER BY date DESC''', (user_id, start_date))
    result = cursor.fetchall()
    conn.close()
    return result

def get_all_timesheet_by_period(start_date, end_date, store=None, show_unconfirmed=False):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    
    confirmed_filter = "" if show_unconfirmed else "AND t.confirmed = 1"
    
    if store:
        cursor.execute(f'''SELECT e.full_name, e.position, e.store, t.date, t.status, t.check_in, t.check_out, t.hours, t.notes, t.confirmed
                          FROM timesheet t 
                          JOIN employees e ON t.user_id = e.user_id
                          WHERE t.date BETWEEN ? AND ? AND e.store = ? {confirmed_filter}
                          ORDER BY e.store, e.full_name, t.date''', (start_date, end_date, store))
    else:
        cursor.execute(f'''SELECT e.full_name, e.position, e.store, t.date, t.status, t.check_in, t.check_out, t.hours, t.notes, t.confirmed
                          FROM timesheet t 
                          JOIN employees e ON t.user_id = e.user_id
                          WHERE t.date BETWEEN ? AND ? {confirmed_filter}
                          ORDER BY e.store, e.full_name, t.date''', (start_date, end_date))
    
    result = cursor.fetchall()
    conn.close()
    return result

# НОВАЯ ФУНКЦИЯ ДЛЯ ПАРСИНГА ДАТ
def parse_date(date_str):
    """Парсинг даты из строки в формате ГГГГ-ММ-ДД"""
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date().isoformat()
    except ValueError:
        return None

# Функции для подтверждения смен
def get_unconfirmed_shifts(store=None):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    
    today = date.today().isoformat()
    
    if store:
        cursor.execute('''SELECT t.id, e.full_name, e.position, e.store, t.date, t.check_in, t.check_out, t.hours, t.notes
                          FROM timesheet t 
                          JOIN employees e ON t.user_id = e.user_id
                          WHERE t.date = ? AND t.confirmed = 0 AND e.store = ?
                          ORDER BY e.full_name''', (today, store))
    else:
        cursor.execute('''SELECT t.id, e.full_name, e.position, e.store, t.date, t.check_in, t.check_out, t.hours, t.notes
                          FROM timesheet t 
                          JOIN employees e ON t.user_id = e.user_id
                          WHERE t.date = ? AND t.confirmed = 0
                          ORDER BY e.store, e.full_name''', (today,))
    
    result = cursor.fetchall()
    conn.close()
    return result

def get_unconfirmed_shifts_by_period(days=7, store=None):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    
    start_date = (date.today() - timedelta(days=days)).isoformat()
    end_date = date.today().isoformat()
    
    if store:
        cursor.execute('''SELECT t.id, e.full_name, e.position, e.store, t.date, t.check_in, t.check_out, t.hours, t.notes
                          FROM timesheet t 
                          JOIN employees e ON t.user_id = e.user_id
                          WHERE t.date BETWEEN ? AND ? AND t.confirmed = 0 AND e.store = ?
                          ORDER BY t.date DESC, e.full_name''', (start_date, end_date, store))
    else:
        cursor.execute('''SELECT t.id, e.full_name, e.position, e.store, t.date, t.check_in, t.check_out, t.hours, t.notes
                          FROM timesheet t 
                          JOIN employees e ON t.user_id = e.user_id
                          WHERE t.date BETWEEN ? AND ? AND t.confirmed = 0
                          ORDER BY t.date DESC, e.store, e.full_name''', (start_date, end_date))
    
    result = cursor.fetchall()
    conn.close()
    return result

def confirm_shift(shift_id):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    cursor.execute('UPDATE timesheet SET confirmed = 1 WHERE id = ?', (shift_id,))
    conn.commit()
    conn.close()

def confirm_all_shifts(store=None, date_str=None):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    
    if not date_str:
        date_str = date.today().isoformat()
    
    if store:
        cursor.execute('''UPDATE timesheet 
                          SET confirmed = 1 
                          WHERE date = ? AND user_id IN 
                          (SELECT user_id FROM employees WHERE store = ?)''', (date_str, store))
    else:
        cursor.execute('UPDATE timesheet SET confirmed = 1 WHERE date = ?', (date_str,))
    
    conn.commit()
    conn.close()

def get_shift_stats(store=None):
    conn = sqlite3.connect('timesheet.db')
    cursor = conn.cursor()
    
    today = date.today().isoformat()
    
    if store:
        cursor.execute('''SELECT 
                          COUNT(CASE WHEN t.confirmed = 0 THEN 1 END) as unconfirmed,
                          COUNT(CASE WHEN t.confirmed = 1 THEN 1 END) as confirmed,
                          COUNT(*) as total
                          FROM timesheet t 
                          JOIN employees e ON t.user_id = e.user_id
                          WHERE t.date = ? AND e.store = ?''', (today, store))
    else:
        cursor.execute('''SELECT 
                          COUNT(CASE WHEN confirmed = 0 THEN 1 END) as unconfirmed,
                          COUNT(CASE WHEN confirmed = 1 THEN 1 END) as confirmed,
                          COUNT(*) as total
                          FROM timesheet 
                          WHERE date = ?''', (today,))
    
    result = cursor.fetchone()
    conn.close()
    return result

# Обработчики команд
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    employee = get_employee(user_id)
    
    if employee:
        admin_star = " 👑" if employee[5] == 1 else ""
        super_admin_star = " ⭐" if len(employee) > 6 and employee[6] == 1 else ""
        store_info = f"\n🏪 Магазин: {employee[3]}" if employee[3] else ""
        
        keyboard = []
        if not employee[5] and not employee[6]:
            keyboard.append([InlineKeyboardButton("👑 Стать администратором", callback_data="become_admin")])
        
        reply_markup = InlineKeyboardMarkup(keyboard) if keyboard else None
        
        await update.message.reply_text(
            f"👋 С возвращением, {employee[1]}{admin_star}{super_admin_star}!\n"
            f"📌 Должность: {employee[2]}{store_info}\n\n"
            "📋 Доступные команды:\n"
            "/checkin - Начать рабочий день\n"
            "/checkout - Закончить рабочий день\n"
            "/timesheet - Мой табель\n"
            "/stats - Моя статистика\n"
            "/help - Помощь",
            reply_markup=reply_markup
        )
    else:
        keyboard = [[InlineKeyboardButton("📝 Зарегистрироваться", callback_data="register")]]
        await update.message.reply_text(
            "Добро пожаловать! Для работы необходимо зарегистрироваться.",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

async def become_admin_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    employee = get_employee(user_id)
    
    if not employee:
        await query.edit_message_text("❌ Сначала зарегистрируйтесь через /start")
        return
    
    if employee[5] == 1:
        await query.edit_message_text("❌ Вы уже являетесь администратором!")
        return
    
    if employee[6] == 1:
        await query.edit_message_text("❌ Вы являетесь супер-администратором!")
        return
    
    if has_pending_admin_request(user_id):
        await query.edit_message_text(
            "⏳ У вас уже есть ожидающая заявка на становление администратором.\n"
            "Пожалуйста, дождитесь решения супер-администратора."
        )
        return
    
    context.user_data['admin_request'] = {
        'user_id': user_id,
        'user_name': employee[1],
        'user_position': employee[2],
        'user_store': employee[3]
    }
    
    keyboard = [
        [InlineKeyboardButton("✅ Отправить заявку", callback_data="confirm_become_admin")],
        [InlineKeyboardButton("❌ Отмена", callback_data="cancel_become_admin")]
    ]
    
    await query.edit_message_text(
        f"👑 *Запрос на становление администратором*\n\n"
        f"Вы уверены, что хотите отправить заявку?\n\n"
        f"👤 Имя: {employee[1]}\n"
        f"📌 Должность: {employee[2]}\n"
        f"🏪 Магазин: {employee[3]}\n\n"
        f"Заявка будет отправлена супер-администратору на рассмотрение.",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )
    return BECOME_ADMIN_REQUEST

async def confirm_become_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_data = context.user_data.get('admin_request', {})
    
    if not user_data:
        await query.edit_message_text("❌ Ошибка: данные не найдены")
        return ConversationHandler.END
    
    request_id = create_admin_request(
        user_id=user_data['user_id'],
        user_name=user_data['user_name'],
        user_position=user_data['user_position'],
        user_store=user_data['user_store']
    )
    
    super_admin_id = get_super_admin()
    if super_admin_id:
        try:
            keyboard = [
                [InlineKeyboardButton("✅ Одобрить", callback_data=f"approve_admin_{request_id}")],
                [InlineKeyboardButton("❌ Отклонить", callback_data=f"reject_admin_{request_id}")]
            ]
            
            await context.bot.send_message(
                chat_id=super_admin_id,
                text=f"⭐ *Новая заявка на администратора*\n\n"
                     f"👤 Пользователь: {user_data['user_name']}\n"
                     f"📌 Должность: {user_data['user_position']}\n"
                     f"🏪 Магазин: {user_data['user_store']}\n"
                     f"🆔 ID: {user_data['user_id']}\n\n"
                     f"Подтвердите или отклоните заявку:",
                reply_markup=InlineKeyboardMarkup(keyboard),
                parse_mode='Markdown'
            )
        except Exception as e:
            logger.error(f"Не удалось отправить уведомление супер-админу: {e}")
    
    await query.edit_message_text(
        "✅ Заявка на становление администратором успешно отправлена!\n"
        "Супер-администратор рассмотрит вашу заявку в ближайшее время.",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ В главное меню", callback_data="back_to_start")
        ]])
    )
    return ConversationHandler.END

async def cancel_become_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    await query.edit_message_text(
        "❌ Запрос на становление администратором отменен.",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ Назад", callback_data="back_to_start")
        ]])
    )
    return ConversationHandler.END

async def back_to_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    employee = get_employee(user_id)
    
    if employee:
        admin_star = " 👑" if employee[5] == 1 else ""
        super_admin_star = " ⭐" if len(employee) > 6 and employee[6] == 1 else ""
        store_info = f"\n🏪 Магазин: {employee[3]}" if employee[3] else ""
        
        keyboard = []
        if not employee[5] and not employee[6]:
            keyboard.append([InlineKeyboardButton("👑 Стать администратором", callback_data="become_admin")])
        
        reply_markup = InlineKeyboardMarkup(keyboard) if keyboard else None
        
        await query.edit_message_text(
            f"👋 С возвращением, {employee[1]}{admin_star}{super_admin_star}!\n"
            f"📌 Должность: {employee[2]}{store_info}\n\n"
            "📋 Доступные команды:\n"
            "/checkin - Начать рабочий день\n"
            "/checkout - Закончить рабочий день\n"
            "/timesheet - Мой табель\n"
            "/stats - Моя статистика\n"
            "/help - Помощь",
            reply_markup=reply_markup
        )
    else:
        await query.edit_message_text("Добро пожаловать! Используйте /start для начала работы.")

async def approve_admin_request(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if not is_super_admin(query.from_user.id):
        await query.edit_message_text("❌ Только супер-администратор может одобрять заявки")
        return
    
    request_id = int(query.data.replace('approve_admin_', ''))
    request = get_admin_request(request_id)
    
    if not request:
        await query.edit_message_text("❌ Заявка не найдена")
        return
    
    if request[6] != 'pending':
        await query.edit_message_text(f"❌ Заявка уже {request[6]}")
        return
    
    user_id = request[2]
    add_admin(user_id)
    update_admin_request_status(request_id, 'approved')
    
    try:
        await context.bot.send_message(
            chat_id=user_id,
            text="✅ Поздравляем! Ваша заявка на становление администратором одобрена!\n"
                 "Теперь вам доступны все команды администратора."
        )
    except Exception as e:
        logger.error(f"Не удалось уведомить пользователя: {e}")
    
    await query.edit_message_text(
        f"✅ Заявка #{request_id} одобрена!\n"
        f"Пользователь {request[3]} теперь администратор."
    )

async def reject_admin_request(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if not is_super_admin(query.from_user.id):
        await query.edit_message_text("❌ Только супер-администратор может отклонять заявки")
        return
    
    request_id = int(query.data.replace('reject_admin_', ''))
    request = get_admin_request(request_id)
    
    if not request:
        await query.edit_message_text("❌ Заявка не найдена")
        return
    
    if request[6] != 'pending':
        await query.edit_message_text(f"❌ Заявка уже {request[6]}")
        return
    
    update_admin_request_status(request_id, 'rejected')
    
    try:
        await context.bot.send_message(
            chat_id=request[2],
            text="❌ К сожалению, ваша заявка на становление администратором отклонена.\n"
                 "Вы можете попробовать подать заявку позже."
        )
    except Exception as e:
        logger.error(f"Не удалось уведомить пользователя: {e}")
    
    await query.edit_message_text(
        f"❌ Заявка #{request_id} отклонена."
    )

async def show_admin_requests(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_super_admin(user_id):
        await update.message.reply_text("❌ Только для супер-администратора")
        return
    
    requests = get_pending_admin_requests()
    
    if not requests:
        await update.message.reply_text("📋 Нет ожидающих заявок на администратора.")
        return
    
    for req in requests:
        request_id = req[0]
        date_str = datetime.fromisoformat(req[1]).strftime('%d.%m.%Y %H:%M')
        user_name = req[3]
        user_position = req[4]
        user_store = req[5]
        
        keyboard = [
            [InlineKeyboardButton("✅ Одобрить", callback_data=f"approve_admin_{request_id}")],
            [InlineKeyboardButton("❌ Отклонить", callback_data=f"reject_admin_{request_id}")]
        ]
        
        await update.message.reply_text(
            f"📋 *Заявка на администратора #{request_id}*\n"
            f"📅 {date_str}\n"
            f"👤 Имя: {user_name}\n"
            f"📌 Должность: {user_position}\n"
            f"🏪 Магазин: {user_store}",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='Markdown'
        )

async def assign_super_admin_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if not is_super_admin(query.from_user.id):
        await query.edit_message_text("❌ Только супер-администратор может назначать супер-админов")
        return
    
    keyboard = [
        [InlineKeyboardButton("⭐ Назначить супер-администратора", callback_data="assign_super_admin_list")],
        [InlineKeyboardButton("📋 Список супер-админов", callback_data="list_super_admins")],
        [InlineKeyboardButton("◀️ Назад в админку", callback_data="back_to_admin")]
    ]
    
    await query.edit_message_text(
        "⭐ *Управление супер-администраторами*\nВыберите действие:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )

async def list_super_admins(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    super_admins = get_all_super_admins()
    
    if not super_admins:
        await query.edit_message_text(
            "📋 Список супер-администраторов пуст.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Назад", callback_data="assign_super_admin_menu")
            ]])
        )
        return
    
    msg = "⭐ *Список супер-администраторов:*\n\n"
    for i, sa in enumerate(super_admins, 1):
        msg += f"{i}. {sa[1]} ({sa[2]}, {sa[3]})\n"
        msg += f"   🆔 ID: {sa[0]}\n\n"
    
    await query.edit_message_text(
        msg,
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ Назад", callback_data="assign_super_admin_menu")
        ]]),
        parse_mode='Markdown'
    )

async def assign_super_admin_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    admins = get_all_admins()
    
    if not admins:
        await query.edit_message_text(
            "❌ Нет администраторов для назначения супер-администратором.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Назад", callback_data="assign_super_admin_menu")
            ]])
        )
        return
    
    by_store = {}
    for a in admins:
        store = a[3] or "Без магазина"
        if store not in by_store:
            by_store[store] = []
        by_store[store].append(a)
    
    msg = "👥 *Выберите администратора для назначения супер-администратором:*\n\n"
    keyboard = []
    
    for store, admins_list in by_store.items():
        for a in admins_list:
            entries_count = get_employee_stats(a[0])
            button_text = f"{a[1]} ({a[2]}) - {entries_count} записей"
            keyboard.append([InlineKeyboardButton(
                button_text[:40], 
                callback_data=f"select_super_admin_{a[0]}"
            )])
    
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="assign_super_admin_menu")])
    
    await query.edit_message_text(
        msg,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )
    return ASSIGN_SUPER_ADMIN_SELECT

async def select_super_admin_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    target_id = int(query.data.replace('select_super_admin_', ''))
    target = get_employee(target_id)
    
    if not target:
        await query.edit_message_text("❌ Пользователь не найден")
        return ConversationHandler.END
    
    context.user_data['new_super_admin_id'] = target_id
    
    keyboard = [
        [InlineKeyboardButton("✅ Да, назначить", callback_data="confirm_assign_super_admin")],
        [InlineKeyboardButton("❌ Нет, отмена", callback_data="assign_super_admin_list")]
    ]
    
    await query.edit_message_text(
        f"⭐ *Подтверждение назначения*\n\n"
        f"Вы уверены, что хотите назначить супер-администратором?\n\n"
        f"👤 Имя: {target[1]}\n"
        f"📌 Должность: {target[2]}\n"
        f"🏪 Магазин: {target[3]}\n\n"
        f"Этот пользователь получит все права супер-администратора!",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )
    return ASSIGN_SUPER_ADMIN_CONFIRM

async def confirm_assign_super_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    target_id = context.user_data.get('new_super_admin_id')
    if not target_id:
        await query.edit_message_text("❌ Ошибка: пользователь не выбран")
        return ConversationHandler.END
    
    target = get_employee(target_id)
    if not target:
        await query.edit_message_text("❌ Пользователь не найден")
        return ConversationHandler.END
    
    assign_super_admin(target_id)
    
    try:
        await context.bot.send_message(
            chat_id=target_id,
            text="⭐ *Поздравляем!*\n\n"
                 "Вы назначены супер-администратором!\n"
                 "Теперь вам доступны все функции управления ботом, включая:\n"
                 "• Подтверждение заявок на администраторов\n"
                 "• Подтверждение запросов на удаление\n"
                 "• Назначение супер-администраторов\n"
                 "• Управление должностями и магазинами",
            parse_mode='Markdown'
        )
    except Exception as e:
        logger.error(f"Не удалось уведомить пользователя: {e}")
    
    await query.edit_message_text(
        f"✅ Пользователь {target[1]} успешно назначен супер-администратором!",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ В меню супер-админов", callback_data="assign_super_admin_menu")
        ]])
    )
    return ConversationHandler.END

# НОВЫЕ ФУНКЦИИ ДЛЯ ВЫБОРА ПЕРИОДА
async def period_selection_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Меню выбора периода для отчетов"""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(query.from_user.id):
        await query.edit_message_text("❌ Только для администраторов")
        return
    
    keyboard = [
        [InlineKeyboardButton("📅 Последние 7 дней", callback_data="period_7")],
        [InlineKeyboardButton("📅 Последние 14 дней", callback_data="period_14")],
        [InlineKeyboardButton("📅 Последние 30 дней", callback_data="period_30")],
        [InlineKeyboardButton("📅 Последние 90 дней", callback_data="period_90")],
        [InlineKeyboardButton("📅 Весь период", callback_data="period_all")],
        [InlineKeyboardButton("📅 Выбрать даты", callback_data="period_custom")],
        [InlineKeyboardButton("◀️ Назад", callback_data="back_to_admin")]
    ]
    
    await query.edit_message_text(
        "📅 *Выбор периода для отчета*\n"
        "Выберите период или укажите конкретные даты:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )

async def process_period_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора периода"""
    query = update.callback_query
    await query.answer()
    
    period = query.data.replace('period_', '')
    
    end_date = date.today().isoformat()
    
    if period == 'all':
        start_date = '2000-01-01'  # Начало времен
        period_text = "за весь период"
    elif period == 'custom':
        await query.edit_message_text(
            "Введите начальную дату в формате ГГГГ-ММ-ДД\n"
            "Например: 2024-01-01"
        )
        return SELECT_PERIOD_START
    else:
        days = int(period)
        start_date = (date.today() - timedelta(days=days)).isoformat()
        period_text = f"за последние {days} дней"
    
    # Сохраняем период в контекст
    context.user_data['report_start'] = start_date
    context.user_data['report_end'] = end_date
    context.user_data['period_text'] = period_text
    
    # Показываем меню выбора типа экспорта
    keyboard = [
        [InlineKeyboardButton("📥 CSV (только подтвержденные)", callback_data="export_confirmed")],
        [InlineKeyboardButton("📥 CSV (все смены)", callback_data="export_all")],
        [InlineKeyboardButton("📊 Excel (только подтвержденные)", callback_data="excel_confirmed")],
        [InlineKeyboardButton("📊 Excel (все смены)", callback_data="excel_all")],
        [InlineKeyboardButton("◀️ Назад к выбору периода", callback_data="period_selection")]
    ]
    
    await query.edit_message_text(
        f"📅 Выбран период: {period_text}\n\n"
        f"Выберите формат отчета:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return SELECT_PERIOD_TYPE

async def process_custom_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка ввода начальной даты"""
    start_date = parse_date(update.message.text)
    
    if not start_date:
        await update.message.reply_text(
            "❌ Неверный формат даты. Используйте ГГГГ-ММ-ДД\n"
            "Например: 2024-01-01\n\n"
            "Попробуйте снова:"
        )
        return SELECT_PERIOD_START
    
    context.user_data['custom_start'] = start_date
    
    await update.message.reply_text(
        f"✅ Начальная дата: {start_date}\n\n"
        f"Теперь введите конечную дату в формате ГГГГ-ММ-ДД\n"
        f"Например: 2024-01-31"
    )
    return SELECT_PERIOD_END

async def process_custom_end(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка ввода конечной даты"""
    end_date = parse_date(update.message.text)
    
    if not end_date:
        await update.message.reply_text(
            "❌ Неверный формат даты. Используйте ГГГГ-ММ-ДД\n"
            "Например: 2024-01-31\n\n"
            "Попробуйте снова:"
        )
        return SELECT_PERIOD_END
    
    start_date = context.user_data.get('custom_start')
    
    if start_date > end_date:
        await update.message.reply_text(
            "❌ Начальная дата не может быть позже конечной!\n"
            f"Начальная: {start_date}\n"
            f"Конечная: {end_date}\n\n"
            "Введите конечную дату снова:"
        )
        return SELECT_PERIOD_END
    
    context.user_data['report_start'] = start_date
    context.user_data['report_end'] = end_date
    context.user_data['period_text'] = f"с {start_date} по {end_date}"
    
    keyboard = [
        [InlineKeyboardButton("📥 CSV (только подтвержденные)", callback_data="export_confirmed")],
        [InlineKeyboardButton("📥 CSV (все смены)", callback_data="export_all")],
        [InlineKeyboardButton("📊 Excel (только подтвержденные)", callback_data="excel_confirmed")],
        [InlineKeyboardButton("📊 Excel (все смены)", callback_data="excel_all")],
        [InlineKeyboardButton("◀️ Назад к выбору периода", callback_data="period_selection")]
    ]
    
    await update.message.reply_text(
        f"📅 Выбран период: с {start_date} по {end_date}\n\n"
        f"Выберите формат отчета:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return SELECT_PERIOD_TYPE

async def export_with_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспорт с выбранным периодом"""
    query = update.callback_query
    await query.answer()
    
    export_type = query.data
    
    start_date = context.user_data.get('report_start')
    end_date = context.user_data.get('report_end')
    period_text = context.user_data.get('period_text', '')
    
    if not start_date or not end_date:
        await query.edit_message_text("❌ Ошибка: период не выбран")
        return ConversationHandler.END
    
    show_unconfirmed = 'all' in export_type
    is_excel = 'excel' in export_type
    
    entries = get_all_timesheet_by_period(start_date, end_date, show_unconfirmed=show_unconfirmed)
    
    if not entries:
        status_text = "все" if show_unconfirmed else "подтвержденные"
        await query.edit_message_text(f"📊 Нет {status_text} записей за {period_text}")
        return ConversationHandler.END
    
    if is_excel:
        # Экспорт в Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Табель"
        
        headers = ['Сотрудник', 'Должность', 'Магазин', 'Дата', 'Статус', 
                   'Начало', 'Конец', 'Часов', 'Примечание', 'Подтверждено']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row, e in enumerate(entries, 2):
            status_rus = 'Завершен' if e[4] == 'completed' else 'В работе' if e[4] == 'working' else e[4]
            confirmed = 'Да' if e[9] == 1 else 'Нет'
            hours = f"{e[7]:.1f}".replace('.', ',') if e[7] else ''
            
            ws.cell(row=row, column=1, value=e[0])
            ws.cell(row=row, column=2, value=e[1])
            ws.cell(row=row, column=3, value=e[2])
            ws.cell(row=row, column=4, value=e[3])
            ws.cell(row=row, column=5, value=status_rus)
            ws.cell(row=row, column=6, value=e[5] or '')
            ws.cell(row=row, column=7, value=e[6] or '')
            ws.cell(row=row, column=8, value=hours)
            ws.cell(row=row, column=9, value=e[8] or '')
            ws.cell(row=row, column=10, value=confirmed)
        
        for col in range(1, 11):
            column_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            wb.save(tmp_file.name)
            tmp_file.seek(0)
            
            type_text = "все" if show_unconfirmed else "подтвержденные"
            filename = f"timesheet_{type_text}_{start_date}_to_{end_date}.xlsx"
            
            with open(tmp_file.name, 'rb') as f:
                await query.message.reply_document(
                    document=f,
                    filename=filename,
                    caption=f"📊 Табель ({type_text}) {period_text}"
                )
            
            os.unlink(tmp_file.name)
    else:
        # Экспорт в CSV
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        
        writer.writerow(['Сотрудник', 'Должность', 'Магазин', 'Дата', 'Статус', 
                         'Начало', 'Конец', 'Часов', 'Примечание', 'Подтверждено'])
        
        for e in entries:
            status_rus = 'Завершен' if e[4] == 'completed' else 'В работе' if e[4] == 'working' else e[4]
            
            writer.writerow([
                e[0], e[1], e[2], e[3], status_rus, e[5] or '', e[6] or '',
                f"{e[7]:.1f}".replace('.', ',') if e[7] else '', e[8] or '',
                'Да' if e[9] == 1 else 'Нет'
            ])
        
        csv_data = output.getvalue()
        output.close()
        
        type_text = "все" if show_unconfirmed else "подтвержденные"
        filename = f"timesheet_{type_text}_{start_date}_to_{end_date}.csv"
        
        await query.message.reply_document(
            document=csv_data.encode('utf-8-sig'),
            filename=filename,
            caption=f"📊 Табель ({type_text}) {period_text}"
        )
    
    # Очищаем данные периода
    context.user_data.pop('report_start', None)
    context.user_data.pop('report_end', None)
    context.user_data.pop('period_text', None)
    context.user_data.pop('custom_start', None)
    
    # Возвращаемся в меню выбора периода
    await period_selection_menu(update, context)
    return ConversationHandler.END

async def register_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    positions = get_all_positions()
    stores = get_all_stores()
    
    if not positions or not stores:
        await query.edit_message_text(
            "Сначала администратор должен создать список должностей и магазинов.\n"
            "Пожалуйста, обратитесь к администратору."
        )
        return ConversationHandler.END
    
    context.user_data['positions'] = positions
    context.user_data['stores'] = stores
    
    keyboard = []
    for pos in positions:
        keyboard.append([InlineKeyboardButton(pos, callback_data=f"select_pos_{pos}")])
    
    await query.edit_message_text(
        "Выберите вашу должность из списка:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return REGISTER_POSITION

async def select_position(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    position = query.data.replace('select_pos_', '')
    context.user_data['selected_position'] = position
    
    stores = context.user_data.get('stores', [])
    
    keyboard = []
    for store_name, store_address in stores:
        button_text = f"{store_name}" + (f" ({store_address})" if store_address else "")
        keyboard.append([InlineKeyboardButton(button_text, callback_data=f"select_store_{store_name}")])
    
    await query.edit_message_text(
        f"Выбрана должность: {position}\n\nТеперь выберите ваш магазин:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return REGISTER_STORE

async def select_store(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    store = query.data.replace('select_store_', '')
    position = context.user_data.get('selected_position')
    
    await query.edit_message_text("Введите ваше полное имя:")
    context.user_data['selected_store'] = store
    return REGISTER_NAME

async def register_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    full_name = update.message.text
    position = context.user_data.get('selected_position')
    store = context.user_data.get('selected_store')
    
    is_admin, is_super_admin = add_employee(user_id, full_name, position, store)
    
    admin_text = "\n\n👑 Вы первый пользователь, поэтому вы назначены администратором!" if is_admin else ""
    super_admin_text = "\n⭐ Вы также являетесь супер-администратором!" if is_super_admin else ""
    
    await update.message.reply_text(
        f"✅ Регистрация завершена!\n"
        f"Имя: {full_name}\n"
        f"Должность: {position}\n"
        f"Магазин: {store}{admin_text}{super_admin_text}"
    )
    return ConversationHandler.END

async def checkin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not get_employee(user_id):
        await update.message.reply_text("❌ Сначала зарегистрируйтесь через /start")
        return
    
    time = add_checkin(user_id)
    await update.message.reply_text(f"✅ Начало рабочего дня отмечено в {time}")

async def checkout(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not get_employee(user_id):
        await update.message.reply_text("❌ Сначала зарегистрируйтесь через /start")
        return
    
    time, hours = add_checkout(user_id)
    if time:
        await update.message.reply_text(f"✅ Конец рабочего дня отмечен в {time}\n⏱ Отработано часов: {hours:.1f}")
    else:
        await update.message.reply_text("❌ Сначала отметьте начало дня через /checkin")

async def timesheet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    employee = get_employee(user_id)
    if not employee:
        await update.message.reply_text("❌ Сначала зарегистрируйтесь")
        return
    
    if context.args:
        try:
            days = int(context.args[0])
        except:
            days = 7
    else:
        days = 7
    
    entries = get_timesheet(user_id, days)
    if not entries:
        await update.message.reply_text(f"📊 За последние {days} дней записей нет")
        return
    
    msg = f"📋 *Табель {employee[1]} ({employee[2]}, {employee[3]})*\n"
    msg += f"📅 За последние {days} дней:\n\n"
    
    total_hours = 0
    for e in entries:
        date_obj = datetime.strptime(e[2], '%Y-%m-%d').strftime('%d.%m.%Y')
        status = "✅" if e[3] == 'completed' else "⏳"
        hours = f"({e[6]:.1f}ч)" if e[6] else ""
        confirmed = " ✓" if len(e) > 8 and e[8] == 1 else " ⏳"
        msg += f"{date_obj} {status}{confirmed} {e[4]}-{e[5] or '...'} {hours}\n"
        if e[6]:
            total_hours += e[6]
    
    msg += f"\n⏱ Всего часов: {total_hours:.1f}"
    
    await update.message.reply_text(msg, parse_mode='Markdown')

async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    employee = get_employee(user_id)
    if not employee:
        await update.message.reply_text("❌ Сначала зарегистрируйтесь")
        return
    
    entries = get_timesheet(user_id, 30)
    total_hours = sum(e[6] for e in entries if e[6])
    days_worked = len([e for e in entries if e[3] == 'completed'])
    avg_hours = total_hours / days_worked if days_worked > 0 else 0
    
    days_of_week = {0: 'Пн', 1: 'Вт', 2: 'Ср', 3: 'Чт', 4: 'Пт', 5: 'Сб', 6: 'Вс'}
    day_stats = {d: 0 for d in range(7)}
    
    for e in entries:
        if e[6]:
            entry_date = datetime.strptime(e[2], '%Y-%m-%d')
            day_stats[entry_date.weekday()] += 1
    
    msg = f"""
📊 *Статистика за 30 дней*

👤 {employee[1]}
📌 {employee[2]}, 🏪 {employee[3]}

📅 Отработано дней: {days_worked}
⏱ Всего часов: {total_hours:.1f}
📈 Среднее часов: {avg_hours:.1f}

📆 По дням недели:
"""
    for day_num, count in day_stats.items():
        if count > 0:
            msg += f"{days_of_week[day_num]}: {count} дней\n"
    
    await update.message.reply_text(msg, parse_mode='Markdown')

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    admin = is_admin(user_id)
    super_admin = is_super_admin(user_id)
    
    help_text = """
📋 *Команды бота:*

👤 *Для всех:*
/start - Начать работу
/checkin - Начать рабочий день
/checkout - Закончить рабочий день
/timesheet [дней] - Мой табель
/stats - Моя статистика
/help - Помощь

"""
    if not admin and not super_admin:
        help_text += """
👑 *Для получения прав администратора:*
В меню после /start нажмите кнопку "Стать администратором"
"""
    
    if admin:
        help_text += """
👑 *Для администраторов:*
/admin - Панель управления
/employees - Список сотрудников
/export [дней] - Выгрузить табель в CSV (только подтвержденные)
/exportall [дней] - Выгрузить все смены (включая неподтвержденные)
/excel [дней] - Выгрузить табель в Excel (только подтвержденные)
/excelall [дней] - Выгрузить все смены в Excel
/exportdates ГГГГ-ММ-ДД ГГГГ-ММ-ДД - Выгрузить за конкретные даты
/addadmin - Добавить администратора
/stores - Магазины и сотрудники
/confirm - Подтверждение смен
/delete - Запросить удаление
/positions - Управление должностями
/stores_list - Управление магазинами
"""
    
    if super_admin:
        help_text += """
⭐ *Для супер-администратора:*
/requests - Просмотр запросов на удаление
/adminrequests - Просмотр заявок на администратора
/superadmin - Управление супер-администраторами
"""
    
    await update.message.reply_text(help_text, parse_mode='Markdown')

# НОВАЯ КОМАНДА ДЛЯ ЭКСПОРТА ПО ДАТАМ
async def export_by_dates(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспорт за конкретные даты: /exportdates 2024-01-01 2024-01-31"""
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Только для администраторов")
        return
    
    if len(context.args) != 2:
        await update.message.reply_text(
            "❌ Неверный формат. Используйте:\n"
            "/exportdates ГГГГ-ММ-ДД ГГГГ-ММ-ДД\n"
            "Например: /exportdates 2024-01-01 2024-01-31"
        )
        return
    
    start_date = parse_date(context.args[0])
    end_date = parse_date(context.args[1])
    
    if not start_date or not end_date:
        await update.message.reply_text(
            "❌ Неверный формат даты. Используйте ГГГГ-ММ-ДД\n"
            "Например: 2024-01-01"
        )
        return
    
    if start_date > end_date:
        await update.message.reply_text(
            "❌ Начальная дата не может быть позже конечной!"
        )
        return
    
    # Спрашиваем тип экспорта
    context.user_data['export_start'] = start_date
    context.user_data['export_end'] = end_date
    
    keyboard = [
        [InlineKeyboardButton("📥 CSV (только подтвержденные)", callback_data="dates_export_confirmed")],
        [InlineKeyboardButton("📥 CSV (все смены)", callback_data="dates_export_all")],
        [InlineKeyboardButton("📊 Excel (только подтвержденные)", callback_data="dates_excel_confirmed")],
        [InlineKeyboardButton("📊 Excel (все смены)", callback_data="dates_excel_all")]
    ]
    
    await update.message.reply_text(
        f"📅 Период: с {start_date} по {end_date}\n\n"
        f"Выберите формат отчета:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def process_dates_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка экспорта по датам"""
    query = update.callback_query
    await query.answer()
    
    export_type = query.data
    
    start_date = context.user_data.get('export_start')
    end_date = context.user_data.get('export_end')
    
    if not start_date or not end_date:
        await query.edit_message_text("❌ Ошибка: даты не выбраны")
        return
    
    show_unconfirmed = 'all' in export_type
    is_excel = 'excel' in export_type
    
    entries = get_all_timesheet_by_period(start_date, end_date, show_unconfirmed=show_unconfirmed)
    
    if not entries:
        status_text = "все" if show_unconfirmed else "подтвержденные"
        await query.edit_message_text(f"📊 Нет {status_text} записей с {start_date} по {end_date}")
        return
    
    if is_excel:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Табель"
        
        headers = ['Сотрудник', 'Должность', 'Магазин', 'Дата', 'Статус', 
                   'Начало', 'Конец', 'Часов', 'Примечание', 'Подтверждено']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row, e in enumerate(entries, 2):
            status_rus = 'Завершен' if e[4] == 'completed' else 'В работе' if e[4] == 'working' else e[4]
            confirmed = 'Да' if e[9] == 1 else 'Нет'
            hours = f"{e[7]:.1f}".replace('.', ',') if e[7] else ''
            
            ws.cell(row=row, column=1, value=e[0])
            ws.cell(row=row, column=2, value=e[1])
            ws.cell(row=row, column=3, value=e[2])
            ws.cell(row=row, column=4, value=e[3])
            ws.cell(row=row, column=5, value=status_rus)
            ws.cell(row=row, column=6, value=e[5] or '')
            ws.cell(row=row, column=7, value=e[6] or '')
            ws.cell(row=row, column=8, value=hours)
            ws.cell(row=row, column=9, value=e[8] or '')
            ws.cell(row=row, column=10, value=confirmed)
        
        for col in range(1, 11):
            column_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            wb.save(tmp_file.name)
            tmp_file.seek(0)
            
            type_text = "all" if show_unconfirmed else "confirmed"
            filename = f"timesheet_{type_text}_{start_date}_to_{end_date}.xlsx"
            
            with open(tmp_file.name, 'rb') as f:
                await query.message.reply_document(
                    document=f,
                    filename=filename,
                    caption=f"📊 Табель с {start_date} по {end_date}"
                )
            
            os.unlink(tmp_file.name)
    else:
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        
        writer.writerow(['Сотрудник', 'Должность', 'Магазин', 'Дата', 'Статус', 
                         'Начало', 'Конец', 'Часов', 'Примечание', 'Подтверждено'])
        
        for e in entries:
            status_rus = 'Завершен' if e[4] == 'completed' else 'В работе' if e[4] == 'working' else e[4]
            
            writer.writerow([
                e[0], e[1], e[2], e[3], status_rus, e[5] or '', e[6] or '',
                f"{e[7]:.1f}".replace('.', ',') if e[7] else '', e[8] or '',
                'Да' if e[9] == 1 else 'Нет'
            ])
        
        csv_data = output.getvalue()
        output.close()
        
        type_text = "all" if show_unconfirmed else "confirmed"
        filename = f"timesheet_{type_text}_{start_date}_to_{end_date}.csv"
        
        await query.message.reply_document(
            document=csv_data.encode('utf-8-sig'),
            filename=filename,
            caption=f"📊 Табель с {start_date} по {end_date}"
        )
    
    # Очищаем данные
    context.user_data.pop('export_start', None)
    context.user_data.pop('export_end', None)

# Административные функции
async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Только для администраторов")
        return
    
    keyboard = [
        [InlineKeyboardButton("👥 Все сотрудники", callback_data="admin_list")],
        [InlineKeyboardButton("📊 По магазинам", callback_data="admin_by_store")],
        [InlineKeyboardButton("📥 Экспорт CSV (подтв.)", callback_data="admin_export_menu")],
        [InlineKeyboardButton("📥 Экспорт CSV (все)", callback_data="admin_export_all_menu")],
        [InlineKeyboardButton("📊 Экспорт Excel (подтв.)", callback_data="admin_excel_menu")],
        [InlineKeyboardButton("📊 Экспорт Excel (все)", callback_data="admin_excel_all_menu")],
        [InlineKeyboardButton("📅 Выбрать период", callback_data="period_selection")],  # НОВАЯ КНОПКА
        [InlineKeyboardButton("➕ Добавить админа", callback_data="admin_add")],
        [InlineKeyboardButton("📈 Статистика по магазинам", callback_data="admin_store_stats")],
        [InlineKeyboardButton("✅ Подтверждение смен", callback_data="admin_confirm")],
        [InlineKeyboardButton("🗑 Запросить удаление", callback_data="admin_delete_menu")],
        [InlineKeyboardButton("📋 Управление должностями", callback_data="admin_positions_menu")],
        [InlineKeyboardButton("🏪 Управление магазинами", callback_data="admin_stores_menu")]
    ]
    
    if is_super_admin(user_id):
        keyboard.append([InlineKeyboardButton("📋 Запросы на удаление", callback_data="admin_requests")])
        keyboard.append([InlineKeyboardButton("👑 Заявки в админы", callback_data="admin_admin_requests")])
        keyboard.append([InlineKeyboardButton("⭐ Управление супер-админами", callback_data="assign_super_admin_menu")])
    
    await update.message.reply_text(
        "🔐 *Панель администратора*\nВыберите действие:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )

# Функции для управления должностями
async def positions_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    keyboard = [
        [InlineKeyboardButton("➕ Создать должность", callback_data="create_position")],
        [InlineKeyboardButton("📋 Список должностей", callback_data="list_positions")],
        [InlineKeyboardButton("🗑 Удалить должность", callback_data="delete_position_menu")],
        [InlineKeyboardButton("◀️ Назад в админку", callback_data="back_to_admin")]
    ]
    
    await query.edit_message_text(
        "📋 *Управление должностями*\nВыберите действие:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )

async def create_position_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    await query.edit_message_text(
        "Введите название новой должности:"
    )
    return CREATE_POSITION_NAME

async def create_position_save(update: Update, context: ContextTypes.DEFAULT_TYPE):
    position_name = update.message.text.strip()
    user_id = update.effective_user.id
    
    if add_position(position_name, user_id):
        await update.message.reply_text(
            f"✅ Должность '{position_name}' успешно создана!",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ В меню должностей", callback_data="admin_positions_menu")
            ]])
        )
    else:
        await update.message.reply_text(
            f"❌ Должность '{position_name}' уже существует!",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Попробовать снова", callback_data="create_position")
            ]])
        )
    return ConversationHandler.END

async def list_positions(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    positions = get_all_positions()
    
    if not positions:
        await query.edit_message_text(
            "📋 Список должностей пуст.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Назад", callback_data="admin_positions_menu")
            ]])
        )
        return
    
    msg = "📋 *Список должностей:*\n\n"
    for i, pos in enumerate(positions, 1):
        msg += f"{i}. {pos}\n"
    
    await query.edit_message_text(
        msg,
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ Назад", callback_data="admin_positions_menu")
        ]]),
        parse_mode='Markdown'
    )

async def delete_position_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    positions = get_all_positions()
    
    if not positions:
        await query.edit_message_text(
            "📋 Список должностей пуст.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Назад", callback_data="admin_positions_menu")
            ]])
        )
        return
    
    keyboard = []
    for pos in positions:
        keyboard.append([InlineKeyboardButton(f"🗑 {pos}", callback_data=f"delete_position_{pos}")])
    
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="admin_positions_menu")])
    
    await query.edit_message_text(
        "Выберите должность для удаления:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def delete_position_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    position = query.data.replace('delete_position_', '')
    
    employees = get_all_employees()
    used_by = [e for e in employees if e[2] == position]
    
    if used_by:
        msg = f"❌ Нельзя удалить должность '{position}', так как она используется:\n"
        for e in used_by[:5]:
            msg += f"  • {e[1]} ({e[3]})\n"
        if len(used_by) > 5:
            msg += f"  • и еще {len(used_by) - 5} сотрудников\n"
        
        await query.edit_message_text(
            msg,
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Назад", callback_data="delete_position_menu")
            ]])
        )
        return
    
    delete_position(position)
    
    await query.edit_message_text(
        f"✅ Должность '{position}' удалена!",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ В меню должностей", callback_data="admin_positions_menu")
        ]])
    )

# Функции для управления магазинами
async def stores_management_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    keyboard = [
        [InlineKeyboardButton("➕ Создать магазин", callback_data="create_store")],
        [InlineKeyboardButton("📋 Список магазинов", callback_data="list_stores")],
        [InlineKeyboardButton("🗑 Удалить магазин", callback_data="delete_store_from_list_menu")],
        [InlineKeyboardButton("◀️ Назад в админку", callback_data="back_to_admin")]
    ]
    
    await query.edit_message_text(
        "🏪 *Управление магазинами*\nВыберите действие:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )

async def create_store_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    await query.edit_message_text(
        "Введите название нового магазина:"
    )
    return CREATE_STORE_NAME

async def create_store_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    store_name = update.message.text.strip()
    context.user_data['new_store_name'] = store_name
    
    await update.message.reply_text(
        f"Введите адрес магазина '{store_name}':\n"
        "(или отправьте '-' если адрес не нужен)"
    )
    return CREATE_STORE_ADDRESS

async def create_store_save(update: Update, context: ContextTypes.DEFAULT_TYPE):
    address = update.message.text.strip()
    store_name = context.user_data.get('new_store_name')
    user_id = update.effective_user.id
    
    if address == '-':
        address = ''
    
    if add_store(store_name, address, user_id):
        await update.message.reply_text(
            f"✅ Магазин '{store_name}' успешно создан!",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ В меню магазинов", callback_data="admin_stores_menu")
            ]])
        )
    else:
        await update.message.reply_text(
            f"❌ Магазин '{store_name}' уже существует!",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Попробовать снова", callback_data="create_store")
            ]])
        )
    return ConversationHandler.END

async def list_stores(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    stores = get_all_stores()
    
    if not stores:
        await query.edit_message_text(
            "📋 Список магазинов пуст.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Назад", callback_data="admin_stores_menu")
            ]])
        )
        return
    
    msg = "🏪 *Список магазинов:*\n\n"
    for i, (store_name, store_address) in enumerate(stores, 1):
        msg += f"{i}. *{store_name}*\n"
        if store_address:
            msg += f"   📍 {store_address}\n"
        msg += "\n"
    
    await query.edit_message_text(
        msg,
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ Назад", callback_data="admin_stores_menu")
        ]]),
        parse_mode='Markdown'
    )

async def delete_store_from_list_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    stores = get_all_stores()
    
    if not stores:
        await query.edit_message_text(
            "📋 Список магазинов пуст.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Назад", callback_data="admin_stores_menu")
            ]])
        )
        return
    
    keyboard = []
    for store_name, store_address in stores:
        button_text = f"🗑 {store_name}"
        if store_address:
            button_text += f" ({store_address})"
        keyboard.append([InlineKeyboardButton(button_text[:40], callback_data=f"delete_store_list_{store_name}")])
    
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="admin_stores_menu")])
    
    await query.edit_message_text(
        "Выберите магазин для удаления из списка:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def delete_store_from_list_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    store_name = query.data.replace('delete_store_list_', '')
    
    employees = get_employees_by_store(store_name)
    
    if employees:
        msg = f"❌ Нельзя удалить магазин '{store_name}', так как в нем есть сотрудники:\n"
        for e in employees[:5]:
            msg += f"  • {e[1]} ({e[2]})\n"
        if len(employees) > 5:
            msg += f"  • и еще {len(employees) - 5} сотрудников\n"
        msg += "\nСначала удалите или переведите сотрудников."
        
        await query.edit_message_text(
            msg,
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Назад", callback_data="delete_store_from_list_menu")
            ]])
        )
        return
    
    delete_store_from_list(store_name)
    
    await query.edit_message_text(
        f"✅ Магазин '{store_name}' удален из списка!",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ В меню магазинов", callback_data="admin_stores_menu")
        ]])
    )

# Функция для списка сотрудников
async def employees_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.callback_query:
        query = update.callback_query
        await query.answer()
        user_id = query.from_user.id
        message_func = query.edit_message_text
        is_callback = True
    else:
        user_id = update.effective_user.id
        message_func = update.message.reply_text
        is_callback = False
    
    if not is_admin(user_id):
        await message_func("❌ Только для администраторов")
        return
    
    employees = get_all_employees()
    
    if not employees:
        await message_func("❌ Нет зарегистрированных сотрудников")
        return
    
    by_store = {}
    for e in employees:
        store = e[3] or "Без магазина"
        if store not in by_store:
            by_store[store] = []
        by_store[store].append(e)
    
    msg = "👥 *Все сотрудники*\n\n"
    for store, emps in by_store.items():
        msg += f"🏪 *{store}*\n"
        for e in emps:
            admin = "👑 " if e[5] == 1 else ""
            super_admin = "⭐ " if len(e) > 6 and e[6] == 1 else ""
            msg += f"  {super_admin}{admin}{e[1]} - {e[2]}\n"
        msg += "\n"
    
    if len(msg) > 4000:
        if is_callback:
            await query.edit_message_text(msg[:4000] + "\n\n*Сообщение продолжается...*", parse_mode='Markdown')
            for i in range(4000, len(msg), 4000):
                await context.bot.send_message(
                    chat_id=user_id,
                    text=msg[i:i+4000],
                    parse_mode='Markdown'
                )
        else:
            for i in range(0, len(msg), 4000):
                await update.message.reply_text(msg[i:i+4000], parse_mode='Markdown')
    else:
        await message_func(msg, parse_mode='Markdown')
    
    if is_callback:
        keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="back_to_admin")]]
        await context.bot.send_message(
            chat_id=user_id,
            text="Выберите действие:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

# Функции экспорта (стандартные)
async def export_timesheet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспорт только подтвержденных смен в CSV"""
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Только для администраторов")
        return
    
    if context.args:
        try:
            days = int(context.args[0])
        except:
            days = 30
    else:
        days = 30
    
    end_date = date.today().isoformat()
    start_date = (date.today() - timedelta(days=days)).isoformat()
    
    entries = get_all_timesheet_by_period(start_date, end_date, show_unconfirmed=False)
    
    if not entries:
        await update.message.reply_text(f"📊 Нет подтвержденных записей за последние {days} дней")
        return
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    writer.writerow(['Сотрудник', 'Должность', 'Магазин', 'Дата', 'Статус', 
                     'Начало', 'Конец', 'Часов', 'Примечание', 'Подтверждено'])
    
    for e in entries:
        status_rus = 'Завершен' if e[4] == 'completed' else 'В работе' if e[4] == 'working' else e[4]
        
        writer.writerow([
            e[0], e[1], e[2], e[3], status_rus, e[5] or '', e[6] or '',
            f"{e[7]:.1f}".replace('.', ',') if e[7] else '', e[8] or '',
            'Да' if e[9] == 1 else 'Нет'
        ])
    
    csv_data = output.getvalue()
    output.close()
    
    filename = f"timesheet_confirmed_{start_date}_to_{end_date}.csv"
    await update.message.reply_document(
        document=csv_data.encode('utf-8-sig'),
        filename=filename,
        caption=f"📊 Табель (только подтвержденные) за последние {days} дней"
    )

async def export_all_timesheet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспорт всех смен (включая неподтвержденные) в CSV"""
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Только для администраторов")
        return
    
    if context.args:
        try:
            days = int(context.args[0])
        except:
            days = 30
    else:
        days = 30
    
    end_date = date.today().isoformat()
    start_date = (date.today() - timedelta(days=days)).isoformat()
    
    entries = get_all_timesheet_by_period(start_date, end_date, show_unconfirmed=True)
    
    if not entries:
        await update.message.reply_text(f"📊 Нет записей за последние {days} дней")
        return
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    writer.writerow(['Сотрудник', 'Должность', 'Магазин', 'Дата', 'Статус', 
                     'Начало', 'Конец', 'Часов', 'Примечание', 'Подтверждено'])
    
    for e in entries:
        status_rus = 'Завершен' if e[4] == 'completed' else 'В работе' if e[4] == 'working' else e[4]
        
        writer.writerow([
            e[0], e[1], e[2], e[3], status_rus, e[5] or '', e[6] or '',
            f"{e[7]:.1f}".replace('.', ',') if e[7] else '', e[8] or '',
            'Да' if e[9] == 1 else 'Нет'
        ])
    
    csv_data = output.getvalue()
    output.close()
    
    filename = f"timesheet_all_{start_date}_to_{end_date}.csv"
    await update.message.reply_document(
        document=csv_data.encode('utf-8-sig'),
        filename=filename,
        caption=f"📊 Табель (все смены) за последние {days} дней"
    )

async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспорт только подтвержденных смен в Excel"""
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Только для администраторов")
        return
    
    if context.args:
        try:
            days = int(context.args[0])
        except:
            days = 30
    else:
        days = 30
    
    end_date = date.today().isoformat()
    start_date = (date.today() - timedelta(days=days)).isoformat()
    
    entries = get_all_timesheet_by_period(start_date, end_date, show_unconfirmed=False)
    
    if not entries:
        await update.message.reply_text(f"📊 Нет подтвержденных записей за последние {days} дней")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Табель"
    
    headers = ['Сотрудник', 'Должность', 'Магазин', 'Дата', 'Статус', 
               'Начало', 'Конец', 'Часов', 'Примечание', 'Подтверждено']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, e in enumerate(entries, 2):
        status_rus = 'Завершен' if e[4] == 'completed' else 'В работе' if e[4] == 'working' else e[4]
        confirmed = 'Да' if e[9] == 1 else 'Нет'
        hours = f"{e[7]:.1f}".replace('.', ',') if e[7] else ''
        
        ws.cell(row=row, column=1, value=e[0])
        ws.cell(row=row, column=2, value=e[1])
        ws.cell(row=row, column=3, value=e[2])
        ws.cell(row=row, column=4, value=e[3])
        ws.cell(row=row, column=5, value=status_rus)
        ws.cell(row=row, column=6, value=e[5] or '')
        ws.cell(row=row, column=7, value=e[6] or '')
        ws.cell(row=row, column=8, value=hours)
        ws.cell(row=row, column=9, value=e[8] or '')
        ws.cell(row=row, column=10, value=confirmed)
    
    for col in range(1, 11):
        column_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[column_letter].auto_size = True
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        wb.save(tmp_file.name)
        tmp_file.seek(0)
        
        with open(tmp_file.name, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename=f"timesheet_confirmed_{start_date}_to_{end_date}.xlsx",
                caption=f"📊 Табель (только подтвержденные) за последние {days} дней"
            )
        
        os.unlink(tmp_file.name)

async def export_all_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспорт всех смен (включая неподтвержденные) в Excel"""
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Только для администраторов")
        return
    
    if context.args:
        try:
            days = int(context.args[0])
        except:
            days = 30
    else:
        days = 30
    
    end_date = date.today().isoformat()
    start_date = (date.today() - timedelta(days=days)).isoformat()
    
    entries = get_all_timesheet_by_period(start_date, end_date, show_unconfirmed=True)
    
    if not entries:
        await update.message.reply_text(f"📊 Нет записей за последние {days} дней")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Табель"
    
    headers = ['Сотрудник', 'Должность', 'Магазин', 'Дата', 'Статус', 
               'Начало', 'Конец', 'Часов', 'Примечание', 'Подтверждено']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, e in enumerate(entries, 2):
        status_rus = 'Завершен' if e[4] == 'completed' else 'В работе' if e[4] == 'working' else e[4]
        confirmed = 'Да' if e[9] == 1 else 'Нет'
        hours = f"{e[7]:.1f}".replace('.', ',') if e[7] else ''
        
        ws.cell(row=row, column=1, value=e[0])
        ws.cell(row=row, column=2, value=e[1])
        ws.cell(row=row, column=3, value=e[2])
        ws.cell(row=row, column=4, value=e[3])
        ws.cell(row=row, column=5, value=status_rus)
        ws.cell(row=row, column=6, value=e[5] or '')
        ws.cell(row=row, column=7, value=e[6] or '')
        ws.cell(row=row, column=8, value=hours)
        ws.cell(row=row, column=9, value=e[8] or '')
        ws.cell(row=row, column=10, value=confirmed)
    
    for col in range(1, 11):
        column_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[column_letter].auto_size = True
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        wb.save(tmp_file.name)
        tmp_file.seek(0)
        
        with open(tmp_file.name, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename=f"timesheet_all_{start_date}_to_{end_date}.xlsx",
                caption=f"📊 Табель (все смены) за последние {days} дней"
            )
        
        os.unlink(tmp_file.name)

# Функции для удаления и запросов
async def delete_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    keyboard = [
        [InlineKeyboardButton("👤 Запросить удаление сотрудника", callback_data="delete_employee_menu")],
        [InlineKeyboardButton("🏪 Запросить удаление магазина", callback_data="delete_store_menu")],
        [InlineKeyboardButton("◀️ Назад в админку", callback_data="back_to_admin")]
    ]
    
    await query.edit_message_text(
        "🗑 *Запрос на удаление*\n"
        "Выберите, что хотите удалить.\n"
        "Запрос будет отправлен супер-администратору на подтверждение.",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )

async def delete_employee_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    employees = get_all_employees()
    
    if not employees:
        await query.edit_message_text(
            "❌ Нет сотрудников для удаления",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Назад", callback_data="admin_delete_menu")
            ]])
        )
        return
    
    by_store = {}
    for e in employees:
        store = e[3] or "Без магазина"
        if store not in by_store:
            by_store[store] = []
        by_store[store].append(e)
    
    msg = "👤 *Выберите сотрудника для удаления:*\n\n"
    keyboard = []
    
    for store, emps in by_store.items():
        for e in emps:
            if e[0] != query.from_user.id and not (len(e) > 6 and e[6] == 1):
                entries_count = get_employee_stats(e[0])
                button_text = f"{e[1]} ({e[2]}) - {entries_count} записей"
                keyboard.append([InlineKeyboardButton(
                    button_text[:40], 
                    callback_data=f"request_delete_employee_{e[0]}"
                )])
    
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="admin_delete_menu")])
    
    if len(keyboard) == 1:
        await query.edit_message_text(
            "❌ Нет доступных сотрудников для удаления",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    else:
        await query.edit_message_text(
            msg,
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='Markdown'
        )

async def request_delete_employee(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    target_id = int(query.data.replace('request_delete_employee_', ''))
    target = get_employee(target_id)
    requester = get_employee(query.from_user.id)
    
    if not target or not requester:
        await query.edit_message_text("❌ Ошибка: сотрудник не найден")
        return
    
    entries_count = get_employee_stats(target_id)
    
    request_id = create_delete_request(
        requester_id=query.from_user.id,
        requester_name=requester[1],
        target_type='employee',
        target_id=str(target_id),
        target_name=target[1]
    )
    
    super_admin_id = get_super_admin()
    if super_admin_id:
        try:
            keyboard = [
                [InlineKeyboardButton("✅ Подтвердить", callback_data=f"approve_request_{request_id}")],
                [InlineKeyboardButton("❌ Отклонить", callback_data=f"reject_request_{request_id}")]
            ]
            
            await context.bot.send_message(
                chat_id=super_admin_id,
                text=f"⭐ *Новый запрос на удаление*\n\n"
                     f"От: {requester[1]} ({requester[2]}, {requester[3]})\n"
                     f"Тип: Сотрудник\n"
                     f"Цель: {target[1]} ({target[2]}, {target[3]})\n"
                     f"Записей в табеле: {entries_count}\n\n"
                     f"Подтвердите или отклоните запрос:",
                reply_markup=InlineKeyboardMarkup(keyboard),
                parse_mode='Markdown'
            )
        except Exception as e:
            logger.error(f"Не удалось отправить уведомление супер-админу: {e}")
    
    await query.edit_message_text(
        f"✅ Запрос на удаление сотрудника {target[1]} отправлен супер-администратору на подтверждение.",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ В меню", callback_data="admin_delete_menu")
        ]])
    )

async def delete_store_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    stores = get_all_stores()
    
    if not stores:
        await query.edit_message_text(
            "❌ Нет магазинов для удаления",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Назад", callback_data="admin_delete_menu")
            ]])
        )
        return
    
    msg = "🏪 *Выберите магазин для удаления:*\n\n"
    keyboard = []
    
    for store_name, store_address in stores:
        employees_count, entries_count = get_store_stats(store_name)
        button_text = f"{store_name} - {employees_count} сотр., {entries_count} записей"
        keyboard.append([InlineKeyboardButton(
            button_text[:40], 
            callback_data=f"request_delete_store_{store_name}"
        )])
    
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="admin_delete_menu")])
    
    await query.edit_message_text(
        msg,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )

async def request_delete_store(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    store = query.data.replace('request_delete_store_', '')
    requester = get_employee(query.from_user.id)
    
    employees_count, entries_count = get_store_stats(store)
    
    request_id = create_delete_request(
        requester_id=query.from_user.id,
        requester_name=requester[1],
        target_type='store',
        target_id=store,
        target_name=store
    )
    
    super_admin_id = get_super_admin()
    if super_admin_id:
        try:
            keyboard = [
                [InlineKeyboardButton("✅ Подтвердить", callback_data=f"approve_request_{request_id}")],
                [InlineKeyboardButton("❌ Отклонить", callback_data=f"reject_request_{request_id}")]
            ]
            
            await context.bot.send_message(
                chat_id=super_admin_id,
                text=f"⭐ *Новый запрос на удаление*\n\n"
                     f"От: {requester[1]} ({requester[2]}, {requester[3]})\n"
                     f"Тип: Магазин\n"
                     f"Цель: {store}\n"
                     f"Сотрудников: {employees_count}\n"
                     f"Записей в табеле: {entries_count}\n\n"
                     f"Подтвердите или отклоните запрос:",
                reply_markup=InlineKeyboardMarkup(keyboard),
                parse_mode='Markdown'
            )
        except Exception as e:
            logger.error(f"Не удалось отправить уведомление супер-админу: {e}")
    
    await query.edit_message_text(
        f"✅ Запрос на удаление магазина {store} отправлен супер-администратору на подтверждение.",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ В меню", callback_data="admin_delete_menu")
        ]])
    )

async def show_requests(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_super_admin(user_id):
        await update.message.reply_text("❌ Только для супер-администратора")
        return
    
    requests = get_pending_requests()
    
    if not requests:
        await update.message.reply_text("📋 Нет ожидающих запросов на удаление.")
        return
    
    for req in requests:
        request_id = req[0]
        date_str = datetime.fromisoformat(req[1]).strftime('%d.%m.%Y %H:%M')
        requester = req[3]
        target_type = "👤 Сотрудник" if req[4] == 'employee' else "🏪 Магазин"
        target_name = req[6]
        
        keyboard = [
            [InlineKeyboardButton("✅ Подтвердить", callback_data=f"approve_request_{request_id}")],
            [InlineKeyboardButton("❌ Отклонить", callback_data=f"reject_request_{request_id}")]
        ]
        
        await update.message.reply_text(
            f"📋 *Запрос #{request_id}*\n"
            f"📅 {date_str}\n"
            f"👤 От: {requester}\n"
            f"📌 Тип: {target_type}\n"
            f"🎯 Цель: {target_name}",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='Markdown'
        )

async def approve_request(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if not is_super_admin(query.from_user.id):
        await query.edit_message_text("❌ Только супер-администратор может подтверждать удаление")
        return
    
    request_id = int(query.data.replace('approve_request_', ''))
    request = get_request(request_id)
    
    if not request:
        await query.edit_message_text("❌ Запрос не найден")
        return
    
    if request[7] != 'pending':
        await query.edit_message_text(f"❌ Запрос уже {request[7]}")
        return
    
    try:
        if request[4] == 'employee':
            target_id = int(request[5])
            target = get_employee(target_id)
            if target:
                delete_employee(target_id)
                result_text = f"✅ Сотрудник {target[1]} удален"
            else:
                result_text = f"❌ Сотрудник не найден"
        
        elif request[4] == 'store':
            store = request[5]
            employees_count, entries_count = get_store_stats(store)
            delete_store_with_employees(store)
            result_text = f"✅ Магазин {store} удален (сотрудников: {employees_count}, записей: {entries_count})"
        
        update_request_status(request_id, 'approved')
        
        try:
            await context.bot.send_message(
                chat_id=request[2],
                text=f"✅ Ваш запрос на удаление {request[4]} {request[6]} одобрен супер-администратором."
            )
        except:
            pass
        
        await query.edit_message_text(result_text)
        
    except Exception as e:
        await query.edit_message_text(f"❌ Ошибка при удалении: {e}")

async def reject_request(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if not is_super_admin(query.from_user.id):
        await query.edit_message_text("❌ Только супер-администратор может отклонять запросы")
        return
    
    request_id = int(query.data.replace('reject_request_', ''))
    request = get_request(request_id)
    
    if not request:
        await query.edit_message_text("❌ Запрос не найден")
        return
    
    if request[7] != 'pending':
        await query.edit_message_text(f"❌ Запрос уже {request[7]}")
        return
    
    update_request_status(request_id, 'rejected')
    
    try:
        await context.bot.send_message(
            chat_id=request[2],
            text=f"❌ Ваш запрос на удаление {request[4]} {request[6]} отклонен супер-администратором."
        )
    except:
        pass
    
    await query.edit_message_text(f"❌ Запрос #{request_id} отклонен")

# Функции для подтверждения смен
async def confirm_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.callback_query:
        query = update.callback_query
        await query.answer()
        user_id = query.from_user.id
        message = query.message
        is_callback = True
    else:
        user_id = update.effective_user.id
        message = update.message
        is_callback = False
    
    if not is_admin(user_id):
        await message.reply_text("❌ Только для администраторов")
        return
    
    employee = get_employee(user_id)
    if not employee:
        await message.reply_text("❌ Сотрудник не найден")
        return
        
    store = employee[3]
    
    stats = get_shift_stats(store)
    
    keyboard = [
        [InlineKeyboardButton("📋 Неподтвержденные сегодня", callback_data="confirm_today")],
        [InlineKeyboardButton("📅 Неподтвержденные за период", callback_data="confirm_period")],
        [InlineKeyboardButton("✅ Подтвердить все сегодня", callback_data="confirm_all_today")],
        [InlineKeyboardButton("🏪 По магазинам", callback_data="confirm_by_store")],
        [InlineKeyboardButton("📊 Статистика подтверждений", callback_data="confirm_stats")],
        [InlineKeyboardButton("◀️ Назад в админку", callback_data="back_to_admin")]
    ]
    
    stats_text = f"\n\n📊 *Статистика на сегодня:*\n"
    stats_text += f"✅ Подтверждено: {stats[1] if stats else 0}\n"
    stats_text += f"⏳ Ожидают: {stats[0] if stats else 0}\n"
    stats_text += f"📝 Всего смен: {stats[2] if stats else 0}"
    
    if is_callback:
        await query.edit_message_text(
            f"🔐 *Меню подтверждения смен*\n"
            f"🏪 Ваш магазин: {store}{stats_text}",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='Markdown'
        )
    else:
        await message.reply_text(
            f"🔐 *Меню подтверждения смен*\n"
            f"🏪 Ваш магазин: {store}{stats_text}",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='Markdown'
        )

async def confirm_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    employee = get_employee(user_id)
    store = employee[3]
    
    unconfirmed = get_unconfirmed_shifts(store)
    
    if not unconfirmed:
        await query.edit_message_text(
            "✅ Все смены за сегодня подтверждены!",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Назад", callback_data="back_to_confirm")
            ]])
        )
        return
    
    msg = f"📋 *Неподтвержденные смены на сегодня*\n\n"
    
    keyboard = []
    for shift in unconfirmed:
        msg += f"👤 {shift[1]} ({shift[2]})\n"
        msg += f"🕐 {shift[5] or '??'} - {shift[6] or '??'}"
        if shift[7]:
            msg += f" ({shift[7]:.1f}ч)"
        if shift[8]:
            msg += f"\n📝 {shift[8]}"
        msg += "\n\n"
        keyboard.append([InlineKeyboardButton(
            f"✅ Подтвердить: {shift[1][:20]}", 
            callback_data=f"confirm_shift_{shift[0]}"
        )])
    
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="back_to_confirm")])
    
    await query.edit_message_text(
        msg,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )

async def confirm_shift_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    shift_id = int(query.data.replace('confirm_shift_', ''))
    confirm_shift(shift_id)
    
    await query.edit_message_text(
        "✅ Смена подтверждена!",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ К списку", callback_data="confirm_today")
        ]])
    )

async def confirm_all_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    employee = get_employee(user_id)
    store = employee[3]
    
    confirm_all_shifts(store)
    
    await query.edit_message_text(
        "✅ Все смены за сегодня подтверждены!",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ Назад", callback_data="back_to_confirm")
        ]])
    )

async def confirm_period_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    keyboard = [
        [InlineKeyboardButton("3 дня", callback_data="confirm_period_3")],
        [InlineKeyboardButton("7 дней", callback_data="confirm_period_7")],
        [InlineKeyboardButton("14 дней", callback_data="confirm_period_14")],
        [InlineKeyboardButton("30 дней", callback_data="confirm_period_30")],
        [InlineKeyboardButton("◀️ Назад", callback_data="back_to_confirm")]
    ]
    
    await query.edit_message_text(
        "📅 Выберите период для просмотра неподтвержденных смен:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def confirm_period_shifts(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    days = int(query.data.replace('confirm_period_', ''))
    
    user_id = query.from_user.id
    employee = get_employee(user_id)
    store = employee[3]
    
    unconfirmed = get_unconfirmed_shifts_by_period(days, store)
    
    if not unconfirmed:
        await query.edit_message_text(
            f"✅ Все смены за последние {days} дней подтверждены!",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("◀️ Назад", callback_data="back_to_confirm")
            ]])
        )
        return
    
    by_date = {}
    for shift in unconfirmed:
        if shift[4] not in by_date:
            by_date[shift[4]] = []
        by_date[shift[4]].append(shift)
    
    msg = f"📋 *Неподтвержденные смены за {days} дней*\n\n"
    
    for date_str, shifts in by_date.items():
        msg += f"📅 *{date_str}*\n"
        for shift in shifts:
            msg += f"  👤 {shift[1]} ({shift[2]})\n"
            msg += f"  🕐 {shift[5] or '??'} - {shift[6] or '??'}"
            if shift[7]:
                msg += f" ({shift[7]:.1f}ч)"
            msg += "\n"
        msg += "\n"
    
    await query.edit_message_text(
        msg,
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ Назад", callback_data="back_to_confirm")
        ]]),
        parse_mode='Markdown'
    )

async def confirm_by_store(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    stores = get_all_stores()
    
    keyboard = []
    for store_name, store_address in stores:
        unconfirmed = get_unconfirmed_shifts(store_name)
        if unconfirmed:
            count = len(unconfirmed)
            keyboard.append([InlineKeyboardButton(
                f"🏪 {store_name} ({count})", 
                callback_data=f"confirm_store_{store_name}"
            )])
    
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="back_to_confirm")])
    
    await query.edit_message_text(
        "Выберите магазин для просмотра:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def confirm_store_shifts(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    store = query.data.replace('confirm_store_', '')
    unconfirmed = get_unconfirmed_shifts(store)
    
    msg = f"🏪 *Магазин: {store}*\n"
    msg += f"📋 Неподтвержденных смен: {len(unconfirmed)}\n\n"
    
    for shift in unconfirmed:
        msg += f"👤 {shift[1]} ({shift[2]})\n"
        msg += f"🕐 {shift[5] or '??'} - {shift[6] or '??'}"
        if shift[7]:
            msg += f" ({shift[7]:.1f}ч)"
        msg += "\n\n"
    
    keyboard = [
        [InlineKeyboardButton("✅ Подтвердить все", callback_data=f"confirm_all_store_{store}")],
        [InlineKeyboardButton("◀️ Назад", callback_data="confirm_by_store")]
    ]
    
    await query.edit_message_text(
        msg,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )

async def confirm_all_store(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    store = query.data.replace('confirm_all_store_', '')
    confirm_all_shifts(store)
    
    await query.edit_message_text(
        f"✅ Все смены в магазине {store} подтверждены!",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ Назад", callback_data="confirm_by_store")
        ]])
    )

async def confirm_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    stores = get_all_stores()
    
    msg = "📊 *Статистика подтверждений*\n\n"
    
    for store_name, store_address in stores:
        stats = get_shift_stats(store_name)
        if stats and stats[2] > 0:
            percent = (stats[1] / stats[2] * 100) if stats[2] > 0 else 0
            msg += f"🏪 *{store_name}*\n"
            msg += f"✅ Подтверждено: {stats[1]}\n"
            msg += f"⏳ Ожидают: {stats[0]}\n"
            msg += f"📈 Процент: {percent:.1f}%\n\n"
    
    await query.edit_message_text(
        msg,
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("◀️ Назад", callback_data="back_to_confirm")
        ]]),
        parse_mode='Markdown'
    )

async def back_to_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await confirm_menu(update, context)

# Функции для экспорта по магазинам
async def export_by_store(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Меню экспорта CSV (только подтвержденные) по магазинам"""
    query = update.callback_query
    await query.answer()
    
    stores = get_all_stores()
    
    if not stores:
        await query.edit_message_text("❌ Нет магазинов с сотрудниками")
        return
    
    keyboard = []
    for store_name, store_address in stores:
        keyboard.append([InlineKeyboardButton(f"🏪 {store_name}", callback_data=f"export_store_confirmed_{store_name}")])
    
    keyboard.append([InlineKeyboardButton("📊 Все магазины", callback_data="export_store_confirmed_all")])
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="back_to_admin")])
    
    await query.edit_message_text(
        "Выберите магазин для экспорта (только подтвержденные смены):",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def export_all_by_store(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Меню экспорта CSV (все смены) по магазинам"""
    query = update.callback_query
    await query.answer()
    
    stores = get_all_stores()
    
    if not stores:
        await query.edit_message_text("❌ Нет магазинов с сотрудниками")
        return
    
    keyboard = []
    for store_name, store_address in stores:
        keyboard.append([InlineKeyboardButton(f"🏪 {store_name}", callback_data=f"export_store_all_{store_name}")])
    
    keyboard.append([InlineKeyboardButton("📊 Все магазины", callback_data="export_store_all_all")])
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="back_to_admin")])
    
    await query.edit_message_text(
        "Выберите магазин для экспорта (все смены):",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def export_store_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспорт данных магазина в CSV (только подтвержденные)"""
    query = update.callback_query
    await query.answer()
    
    data = query.data.replace('export_store_confirmed_', '')
    
    if data == 'all':
        store = None
        filename_prefix = "all_stores_confirmed"
        caption_prefix = "Все магазины"
    else:
        store = data
        filename_prefix = f"{store}_confirmed"
        caption_prefix = f"Магазин: {store}"
    
    end_date = date.today().isoformat()
    start_date = (date.today() - timedelta(days=30)).isoformat()
    
    entries = get_all_timesheet_by_period(start_date, end_date, store, show_unconfirmed=False)
    
    if not entries:
        await query.edit_message_text(f"❌ Нет подтвержденных записей за период")
        return
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    writer.writerow(['Сотрудник', 'Должность', 'Магазин', 'Дата', 'Статус', 
                     'Начало', 'Конец', 'Часов', 'Примечание', 'Подтверждено'])
    
    for e in entries:
        status_rus = 'Завершен' if e[4] == 'completed' else 'В работе' if e[4] == 'working' else e[4]
        
        writer.writerow([
            e[0], e[1], e[2], e[3], status_rus, e[5] or '', e[6] or '',
            f"{e[7]:.1f}".replace('.', ',') if e[7] else '', e[8] or '',
            'Да' if e[9] == 1 else 'Нет'
        ])
    
    csv_data = output.getvalue()
    output.close()
    
    filename = f"timesheet_{filename_prefix}_{start_date}_to_{end_date}.csv"
    await query.message.reply_document(
        document=csv_data.encode('utf-8-sig'),
        filename=filename,
        caption=f"📊 {caption_prefix} (только подтвержденные) за 30 дней"
    )
    
    await admin_panel(update, context)

async def export_all_store_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспорт данных магазина в CSV (все смены)"""
    query = update.callback_query
    await query.answer()
    
    data = query.data.replace('export_store_all_', '')
    
    if data == 'all':
        store = None
        filename_prefix = "all_stores_all"
        caption_prefix = "Все магазины"
    else:
        store = data
        filename_prefix = f"{store}_all"
        caption_prefix = f"Магазин: {store}"
    
    end_date = date.today().isoformat()
    start_date = (date.today() - timedelta(days=30)).isoformat()
    
    entries = get_all_timesheet_by_period(start_date, end_date, store, show_unconfirmed=True)
    
    if not entries:
        await query.edit_message_text(f"❌ Нет записей за период")
        return
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    writer.writerow(['Сотрудник', 'Должность', 'Магазин', 'Дата', 'Статус', 
                     'Начало', 'Конец', 'Часов', 'Примечание', 'Подтверждено'])
    
    for e in entries:
        status_rus = 'Завершен' if e[4] == 'completed' else 'В работе' if e[4] == 'working' else e[4]
        
        writer.writerow([
            e[0], e[1], e[2], e[3], status_rus, e[5] or '', e[6] or '',
            f"{e[7]:.1f}".replace('.', ',') if e[7] else '', e[8] or '',
            'Да' if e[9] == 1 else 'Нет'
        ])
    
    csv_data = output.getvalue()
    output.close()
    
    filename = f"timesheet_{filename_prefix}_{start_date}_to_{end_date}.csv"
    await query.message.reply_document(
        document=csv_data.encode('utf-8-sig'),
        filename=filename,
        caption=f"📊 {caption_prefix} (все смены) за 30 дней"
    )
    
    await admin_panel(update, context)

async def excel_by_store(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Меню экспорта Excel (только подтвержденные) по магазинам"""
    query = update.callback_query
    await query.answer()
    
    stores = get_all_stores()
    
    if not stores:
        await query.edit_message_text("❌ Нет магазинов с сотрудниками")
        return
    
    keyboard = []
    for store_name, store_address in stores:
        keyboard.append([InlineKeyboardButton(f"🏪 {store_name}", callback_data=f"excel_store_confirmed_{store_name}")])
    
    keyboard.append([InlineKeyboardButton("📊 Все магазины", callback_data="excel_store_confirmed_all")])
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="back_to_admin")])
    
    await query.edit_message_text(
        "Выберите магазин для экспорта в Excel (только подтвержденные смены):",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def excel_all_by_store(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Меню экспорта Excel (все смены) по магазинам"""
    query = update.callback_query
    await query.answer()
    
    stores = get_all_stores()
    
    if not stores:
        await query.edit_message_text("❌ Нет магазинов с сотрудниками")
        return
    
    keyboard = []
    for store_name, store_address in stores:
        keyboard.append([InlineKeyboardButton(f"🏪 {store_name}", callback_data=f"excel_store_all_{store_name}")])
    
    keyboard.append([InlineKeyboardButton("📊 Все магазины", callback_data="excel_store_all_all")])
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="back_to_admin")])
    
    await query.edit_message_text(
        "Выберите магазин для экспорта в Excel (все смены):",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def excel_store_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспорт данных магазина в Excel (только подтвержденные)"""
    query = update.callback_query
    await query.answer()
    
    data = query.data.replace('excel_store_confirmed_', '')
    
    if data == 'all':
        store = None
        filename_prefix = "all_stores_confirmed"
        caption_prefix = "Все магазины"
    else:
        store = data
        filename_prefix = f"{store}_confirmed"
        caption_prefix = f"Магазин: {store}"
    
    end_date = date.today().isoformat()
    start_date = (date.today() - timedelta(days=30)).isoformat()
    
    entries = get_all_timesheet_by_period(start_date, end_date, store, show_unconfirmed=False)
    
    if not entries:
        await query.edit_message_text(f"❌ Нет подтвержденных записей за период")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Табель"
    
    headers = ['Сотрудник', 'Должность', 'Магазин', 'Дата', 'Статус', 
               'Начало', 'Конец', 'Часов', 'Примечание', 'Подтверждено']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, e in enumerate(entries, 2):
        status_rus = 'Завершен' if e[4] == 'completed' else 'В работе' if e[4] == 'working' else e[4]
        confirmed = 'Да' if e[9] == 1 else 'Нет'
        hours = f"{e[7]:.1f}".replace('.', ',') if e[7] else ''
        
        ws.cell(row=row, column=1, value=e[0])
        ws.cell(row=row, column=2, value=e[1])
        ws.cell(row=row, column=3, value=e[2])
        ws.cell(row=row, column=4, value=e[3])
        ws.cell(row=row, column=5, value=status_rus)
        ws.cell(row=row, column=6, value=e[5] or '')
        ws.cell(row=row, column=7, value=e[6] or '')
        ws.cell(row=row, column=8, value=hours)
        ws.cell(row=row, column=9, value=e[8] or '')
        ws.cell(row=row, column=10, value=confirmed)
    
    for col in range(1, 11):
        column_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[column_letter].auto_size = True
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        wb.save(tmp_file.name)
        tmp_file.seek(0)
        
        with open(tmp_file.name, 'rb') as f:
            await query.message.reply_document(
                document=f,
                filename=f"timesheet_{filename_prefix}_{start_date}_to_{end_date}.xlsx",
                caption=f"📊 {caption_prefix} (только подтвержденные) за 30 дней"
            )
        
        os.unlink(tmp_file.name)
    
    await admin_panel(update, context)

async def excel_all_store_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспорт данных магазина в Excel (все смены)"""
    query = update.callback_query
    await query.answer()
    
    data = query.data.replace('excel_store_all_', '')
    
    if data == 'all':
        store = None
        filename_prefix = "all_stores_all"
        caption_prefix = "Все магазины"
    else:
        store = data
        filename_prefix = f"{store}_all"
        caption_prefix = f"Магазин: {store}"
    
    end_date = date.today().isoformat()
    start_date = (date.today() - timedelta(days=30)).isoformat()
    
    entries = get_all_timesheet_by_period(start_date, end_date, store, show_unconfirmed=True)
    
    if not entries:
        await query.edit_message_text(f"❌ Нет записей за период")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Табель"
    
    headers = ['Сотрудник', 'Должность', 'Магазин', 'Дата', 'Статус', 
               'Начало', 'Конец', 'Часов', 'Примечание', 'Подтверждено']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, e in enumerate(entries, 2):
        status_rus = 'Завершен' if e[4] == 'completed' else 'В работе' if e[4] == 'working' else e[4]
        confirmed = 'Да' if e[9] == 1 else 'Нет'
        hours = f"{e[7]:.1f}".replace('.', ',') if e[7] else ''
        
        ws.cell(row=row, column=1, value=e[0])
        ws.cell(row=row, column=2, value=e[1])
        ws.cell(row=row, column=3, value=e[2])
        ws.cell(row=row, column=4, value=e[3])
        ws.cell(row=row, column=5, value=status_rus)
        ws.cell(row=row, column=6, value=e[5] or '')
        ws.cell(row=row, column=7, value=e[6] or '')
        ws.cell(row=row, column=8, value=hours)
        ws.cell(row=row, column=9, value=e[8] or '')
        ws.cell(row=row, column=10, value=confirmed)
    
    for col in range(1, 11):
        column_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[column_letter].auto_size = True
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        wb.save(tmp_file.name)
        tmp_file.seek(0)
        
        with open(tmp_file.name, 'rb') as f:
            await query.message.reply_document(
                document=f,
                filename=f"timesheet_{filename_prefix}_{start_date}_to_{end_date}.xlsx",
                caption=f"📊 {caption_prefix} (все смены) за 30 дней"
            )
        
        os.unlink(tmp_file.name)
    
    await admin_panel(update, context)

async def add_admin_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    await query.edit_message_text(
        "Введите Telegram ID пользователя, которого хотите сделать администратором:\n\n"
        "ID можно узнать у пользователя - он должен отправить команду /id боту @userinfobot"
    )
    return ADD_ADMIN_ID

async def add_admin_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        new_admin_id = int(update.message.text)
    except:
        await update.message.reply_text("❌ Неверный формат ID. Введите число.")
        return ADD_ADMIN_ID
    
    employee = get_employee(new_admin_id)
    
    if not employee:
        await update.message.reply_text("❌ Пользователь с таким ID не зарегистрирован в боте")
        return ConversationHandler.END
    
    context.user_data['new_admin_id'] = new_admin_id
    
    keyboard = [
        [InlineKeyboardButton("✅ Да", callback_data="confirm_add_admin")],
        [InlineKeyboardButton("❌ Нет", callback_data="cancel_add_admin")]
    ]
    
    await update.message.reply_text(
        f"Сделать администратором:\n"
        f"👤 {employee[1]}\n"
        f"📌 {employee[2]}\n"
        f"🏪 {employee[3]}\n\n"
        f"Подтвердите действие:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ADD_ADMIN_CONFIRM

async def confirm_add_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    new_admin_id = context.user_data.get('new_admin_id')
    if new_admin_id:
        add_admin(new_admin_id)
        employee = get_employee(new_admin_id)
        await query.edit_message_text(
            f"✅ Пользователь {employee[1]} теперь администратор!"
        )
    
    return ConversationHandler.END

async def cancel_add_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("❌ Добавление администратора отменено")
    return ConversationHandler.END

async def store_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    stores = get_all_stores()
    
    if not stores:
        await query.edit_message_text("❌ Нет магазинов с сотрудниками")
        return
    
    end_date = date.today().isoformat()
    start_date = (date.today() - timedelta(days=30)).isoformat()
    
    msg = "📈 *Статистика по магазинам за 30 дней*\n\n"
    
    for store_name, store_address in stores:
        employees = get_employees_by_store(store_name)
        entries = get_all_timesheet_by_period(start_date, end_date, store_name, show_unconfirmed=False)
        
        total_hours = sum(e[7] for e in entries if e[7])
        total_days = len(set([e[3] for e in entries]))
        
        msg += f"🏪 *{store_name}*\n"
        msg += f"👥 Сотрудников: {len(employees)}\n"
        msg += f"⏱ Всего часов: {total_hours:.1f}\n"
        msg += f"📅 Рабочих дней: {total_days}\n\n"
    
    await query.edit_message_text(msg, parse_mode='Markdown')

async def back_to_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await admin_panel(update, context)

async def stores_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Только для администраторов")
        return
    
    stores = get_all_stores()
    
    if not stores:
        await update.message.reply_text("❌ Нет магазинов с сотрудниками")
        return
    
    msg = "🏪 *Магазины и сотрудники*\n\n"
    
    for store_name, store_address in stores:
        employees = get_employees_by_store(store_name)
        msg += f"*{store_name}*"
        if store_address:
            msg += f" - {store_address}"
        msg += f" ({len(employees)} чел.)\n"
        for e in employees:
            admin = "👑 " if e[5] == 1 else ""
            super_admin = "⭐ " if len(e) > 6 and e[6] == 1 else ""
            msg += f"  {super_admin}{admin}{e[1]} - {e[2]}\n"
        msg += "\n"
    
    await update.message.reply_text(msg, parse_mode='Markdown')

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ Отменено")
    return ConversationHandler.END

def main():
    init_database()
    
    app = Application.builder().token(BOT_TOKEN).build()
    
    # Conversation Handlers
    reg_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(register_start, pattern='^register$')],
        states={
            REGISTER_POSITION: [CallbackQueryHandler(select_position, pattern='^select_pos_')],
            REGISTER_STORE: [CallbackQueryHandler(select_store, pattern='^select_store_')],
            REGISTER_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, register_name)],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    add_admin_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(add_admin_start, pattern='^admin_add$')],
        states={
            ADD_ADMIN_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_admin_id)],
            ADD_ADMIN_CONFIRM: [CallbackQueryHandler(confirm_add_admin, pattern='^confirm_add_admin$'),
                               CallbackQueryHandler(cancel_add_admin, pattern='^cancel_add_admin$')],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    become_admin_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(become_admin_start, pattern='^become_admin$')],
        states={
            BECOME_ADMIN_REQUEST: [
                CallbackQueryHandler(confirm_become_admin, pattern='^confirm_become_admin$'),
                CallbackQueryHandler(cancel_become_admin, pattern='^cancel_become_admin$')
            ],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    create_position_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(create_position_start, pattern='^create_position$')],
        states={
            CREATE_POSITION_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, create_position_save)],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    create_store_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(create_store_start, pattern='^create_store$')],
        states={
            CREATE_STORE_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, create_store_name)],
            CREATE_STORE_ADDRESS: [MessageHandler(filters.TEXT & ~filters.COMMAND, create_store_save)],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    assign_super_admin_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(assign_super_admin_list, pattern='^assign_super_admin_list$')],
        states={
            ASSIGN_SUPER_ADMIN_SELECT: [CallbackQueryHandler(select_super_admin_confirm, pattern='^select_super_admin_')],
            ASSIGN_SUPER_ADMIN_CONFIRM: [CallbackQueryHandler(confirm_assign_super_admin, pattern='^confirm_assign_super_admin$')],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    # НОВЫЙ Conversation handler для выбора периода
    period_selection_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(period_selection_menu, pattern='^period_selection$'),
            CallbackQueryHandler(process_period_selection, pattern='^period_'),
            CallbackQueryHandler(export_with_period, pattern='^(export|excel)_(confirmed|all)$')
        ],
        states={
            SELECT_PERIOD_START: [MessageHandler(filters.TEXT & ~filters.COMMAND, process_custom_start)],
            SELECT_PERIOD_END: [MessageHandler(filters.TEXT & ~filters.COMMAND, process_custom_end)],
            SELECT_PERIOD_TYPE: [CallbackQueryHandler(export_with_period, pattern='^(export|excel)_(confirmed|all)$')],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    # Basic commands
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("checkin", checkin))
    app.add_handler(CommandHandler("checkout", checkout))
    app.add_handler(CommandHandler("timesheet", timesheet))
    app.add_handler(CommandHandler("stats", stats))
    app.add_handler(CommandHandler("admin", admin_panel))
    app.add_handler(CommandHandler("employees", employees_list))
    app.add_handler(CommandHandler("export", export_timesheet))
    app.add_handler(CommandHandler("exportall", export_all_timesheet))
    app.add_handler(CommandHandler("excel", export_excel))
    app.add_handler(CommandHandler("excelall", export_all_excel))
    app.add_handler(CommandHandler("exportdates", export_by_dates))
    app.add_handler(CommandHandler("stores", stores_menu))
    app.add_handler(CommandHandler("confirm", confirm_menu))
    app.add_handler(CommandHandler("requests", show_requests))
    app.add_handler(CommandHandler("adminrequests", show_admin_requests))
    app.add_handler(CommandHandler("superadmin", assign_super_admin_menu))
    app.add_handler(CommandHandler("positions", positions_menu))
    app.add_handler(CommandHandler("stores_list", stores_management_menu))
    
    app.add_handler(CallbackQueryHandler(back_to_start, pattern='^back_to_start$'))
    
    # Admin panel callbacks
    app.add_handler(CallbackQueryHandler(back_to_admin, pattern='^back_to_admin$'))
    app.add_handler(CallbackQueryHandler(export_by_store, pattern='^admin_export_menu$'))
    app.add_handler(CallbackQueryHandler(export_all_by_store, pattern='^admin_export_all_menu$'))
    app.add_handler(CallbackQueryHandler(excel_by_store, pattern='^admin_excel_menu$'))
    app.add_handler(CallbackQueryHandler(excel_all_by_store, pattern='^admin_excel_all_menu$'))
    app.add_handler(CallbackQueryHandler(period_selection_menu, pattern='^period_selection$'))
    app.add_handler(CallbackQueryHandler(store_stats, pattern='^admin_store_stats$'))
    app.add_handler(CallbackQueryHandler(employees_list, pattern='^admin_list$'))
    app.add_handler(CallbackQueryHandler(export_store_data, pattern='^export_store_confirmed_'))
    app.add_handler(CallbackQueryHandler(export_all_store_data, pattern='^export_store_all_'))
    app.add_handler(CallbackQueryHandler(excel_store_data, pattern='^excel_store_confirmed_'))
    app.add_handler(CallbackQueryHandler(excel_all_store_data, pattern='^excel_store_all_'))
    app.add_handler(CallbackQueryHandler(confirm_menu, pattern='^admin_confirm$'))
    app.add_handler(CallbackQueryHandler(delete_menu, pattern='^admin_delete_menu$'))
    app.add_handler(CallbackQueryHandler(delete_employee_menu, pattern='^delete_employee_menu$'))
    app.add_handler(CallbackQueryHandler(delete_store_menu, pattern='^delete_store_menu$'))
    app.add_handler(CallbackQueryHandler(show_requests, pattern='^admin_requests$'))
    app.add_handler(CallbackQueryHandler(show_admin_requests, pattern='^admin_admin_requests$'))
    app.add_handler(CallbackQueryHandler(assign_super_admin_menu, pattern='^assign_super_admin_menu$'))
    app.add_handler(CallbackQueryHandler(list_super_admins, pattern='^list_super_admins$'))
    
    # Date export callbacks
    app.add_handler(CallbackQueryHandler(process_dates_export, pattern='^dates_'))
    
    # Position management callbacks
    app.add_handler(CallbackQueryHandler(positions_menu, pattern='^admin_positions_menu$'))
    app.add_handler(CallbackQueryHandler(list_positions, pattern='^list_positions$'))
    app.add_handler(CallbackQueryHandler(delete_position_menu, pattern='^delete_position_menu$'))
    app.add_handler(CallbackQueryHandler(delete_position_confirm, pattern='^delete_position_'))
    
    # Store management callbacks
    app.add_handler(CallbackQueryHandler(stores_management_menu, pattern='^admin_stores_menu$'))
    app.add_handler(CallbackQueryHandler(list_stores, pattern='^list_stores$'))
    app.add_handler(CallbackQueryHandler(delete_store_from_list_menu, pattern='^delete_store_from_list_menu$'))
    app.add_handler(CallbackQueryHandler(delete_store_from_list_confirm, pattern='^delete_store_list_'))
    
    # Delete request callbacks
    app.add_handler(CallbackQueryHandler(request_delete_employee, pattern='^request_delete_employee_'))
    app.add_handler(CallbackQueryHandler(request_delete_store, pattern='^request_delete_store_'))
    app.add_handler(CallbackQueryHandler(approve_request, pattern='^approve_request_'))
    app.add_handler(CallbackQueryHandler(reject_request, pattern='^reject_request_'))
    
    app.add_handler(CallbackQueryHandler(approve_admin_request, pattern='^approve_admin_'))
    app.add_handler(CallbackQueryHandler(reject_admin_request, pattern='^reject_admin_'))
    
    # Confirmation menu callbacks
    app.add_handler(CallbackQueryHandler(confirm_today, pattern='^confirm_today$'))
    app.add_handler(CallbackQueryHandler(confirm_period_menu, pattern='^confirm_period$'))
    app.add_handler(CallbackQueryHandler(confirm_all_today, pattern='^confirm_all_today$'))
    app.add_handler(CallbackQueryHandler(confirm_by_store, pattern='^confirm_by_store$'))
    app.add_handler(CallbackQueryHandler(confirm_stats, pattern='^confirm_stats$'))
    app.add_handler(CallbackQueryHandler(back_to_confirm, pattern='^back_to_confirm$'))
    app.add_handler(CallbackQueryHandler(confirm_period_shifts, pattern='^confirm_period_\\d+$'))
    app.add_handler(CallbackQueryHandler(confirm_store_shifts, pattern='^confirm_store_'))
    app.add_handler(CallbackQueryHandler(confirm_all_store, pattern='^confirm_all_store_'))
    app.add_handler(CallbackQueryHandler(confirm_shift_action, pattern='^confirm_shift_\\d+$'))
    
    # Conversation handlers
    app.add_handler(reg_conv)
    app.add_handler(add_admin_conv)
    app.add_handler(become_admin_conv)
    app.add_handler(create_position_conv)
    app.add_handler(create_store_conv)
    app.add_handler(assign_super_admin_conv)
    app.add_handler(period_selection_conv)  # Новый conversation handler
    
    print("✅ Бот успешно запущен!")
    app.run_polling()

if __name__ == '__main__':
    main()
